import hashlib
import hmac
import json
import re
import math
import secrets
import time
import zipfile
from copy import copy
from collections import Counter, defaultdict
from functools import wraps
from io import BytesIO
from pathlib import Path

import pandas as pd
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import and_, func, or_

BASE_DIR = Path(__file__).resolve().parent.parent
DB_PATH = BASE_DIR / "data.db"
SOURCE_FILE = BASE_DIR / "isuzu_data.xlsx"
AUTH_FILE = BASE_DIR / "auth.json"
ARRIVAL_DIR = BASE_DIR / "到货文件"
ARRIVAL_COLUMNS = ["合同号", "序号", "零件号", "互换零件号", "零件名", "个数", "单价", "总价"]
ORDER_UPLOAD_COLUMNS = ["序号", "零件号", "互换零件号", "零件名", "中文零件名", "单价", "个数", "总价", "合同号"]
ARRIVAL_MATCH_CACHE = {"signature": None, "index": defaultdict(list)}

# ── Auth ──────────────────────────────────────────────────────────────────────
_AUTH_CONFIG: dict = {}
_FAILED_ATTEMPTS: dict = {}   # {ip: [timestamp, ...]}


def _hash_code(code: str, salt: str, iterations: int) -> str:
    return hashlib.pbkdf2_hmac(
        "sha256", code.encode(), salt.encode(), iterations
    ).hex()


def _setup_auth() -> None:
    global _AUTH_CONFIG
    if AUTH_FILE.exists():
        with open(AUTH_FILE) as f:
            _AUTH_CONFIG = json.load(f)
        return

    code = secrets.token_hex(16)          # 32 hex characters
    salt = secrets.token_hex(16)
    iterations = 600_000
    code_hash = _hash_code(code, salt, iterations)
    _AUTH_CONFIG = {"code_hash": code_hash, "salt": salt, "iterations": iterations}
    with open(AUTH_FILE, "w") as f:
        json.dump(_AUTH_CONFIG, f)

    border = "=" * 54
    print(f"\n{border}")
    print(f"  首次启动，已生成 32 位登录码：")
    print(f"  {code}")
    print(f"  已加盐哈希保存到 auth.json（请妥善保管登录码原文）")
    print(f"{border}\n")


def _verify_code(code: str) -> bool:
    try:
        cfg = _AUTH_CONFIG
        expected = bytes.fromhex(cfg["code_hash"])
        computed = bytes.fromhex(_hash_code(code, cfg["salt"], cfg["iterations"]))
        return hmac.compare_digest(expected, computed)
    except Exception:
        return False


def _is_rate_limited(ip: str) -> bool:
    cutoff = time.time() - 900          # 15-minute window
    attempts = [t for t in _FAILED_ATTEMPTS.get(ip, []) if t > cutoff]
    _FAILED_ATTEMPTS[ip] = attempts
    return len(attempts) >= 5


def _record_failed(ip: str) -> None:
    _FAILED_ATTEMPTS.setdefault(ip, []).append(time.time())


def _is_valid_session(token: str) -> bool:
    session_row = db.session.get(AuthSession, token)
    if session_row is None:
        return False
    if time.time() > session_row.expires_at:
        db.session.delete(session_row)
        db.session.commit()
        return False
    return True


def _clean_sessions() -> None:
    now = time.time()
    db.session.query(AuthSession).filter(AuthSession.expires_at < now).delete(synchronize_session=False)
# ── End Auth ──────────────────────────────────────────────────────────────────


db = SQLAlchemy()


class Order(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    contract_no = db.Column(db.String(120), index=True)
    customer_code = db.Column(db.String(80), index=True)
    sequence = db.Column(db.String(40))
    part_no = db.Column(db.String(80), index=True)
    interchange_part_no = db.Column(db.String(80), index=True)
    part_name = db.Column(db.String(255))
    part_name_cn = db.Column(db.String(255))
    quantity = db.Column(db.Float)
    unit_price = db.Column(db.Float)
    total_price = db.Column(db.Float)


class ArrivalOrder(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    source_file = db.Column(db.String(255), index=True)
    arrival_date = db.Column(db.String(20), index=True)
    contract_no = db.Column(db.String(120), index=True)
    customer_code = db.Column(db.String(80), index=True)
    sequence = db.Column(db.String(40))
    part_no = db.Column(db.String(80), index=True)
    interchange_part_no = db.Column(db.String(80), index=True)
    part_name = db.Column(db.String(255))
    quantity = db.Column(db.Float)
    unit_price = db.Column(db.Float)
    total_price = db.Column(db.Float)


class ContractSummary(db.Model):
    __tablename__ = "contract_summary"
    id = db.Column(db.Integer, primary_key=True)
    contract_no = db.Column(db.String(120), index=True)
    customer_code = db.Column(db.String(80), index=True)
    order_year = db.Column(db.String(10), index=True)
    order_date = db.Column(db.String(10))
    part_item_count = db.Column(db.Integer)
    total_quantity = db.Column(db.Float)
    total_amount = db.Column(db.Float)
    arrival_quantity = db.Column(db.Float)
    arrival_amount = db.Column(db.Float)
    not_arrival_amount = db.Column(db.Float)
    arrival_ratio = db.Column(db.Float)
    arrival_status = db.Column(db.String(20))
    arrival_history_json = db.Column(db.Text)


class CustomerSummary(db.Model):
    __tablename__ = "customer_summary"
    id = db.Column(db.Integer, primary_key=True)
    customer_code = db.Column(db.String(80), unique=True, index=True)
    contract_count = db.Column(db.Integer)
    order_count = db.Column(db.Integer)
    total_amount = db.Column(db.Float)
    latest_contract_time = db.Column(db.String(20))
    latest_contract_no = db.Column(db.String(120))


class PartSummary(db.Model):
    __tablename__ = "part_summary"
    id = db.Column(db.Integer, primary_key=True)
    part_no = db.Column(db.String(80), index=True)
    interchange_part_no = db.Column(db.String(80))
    part_name = db.Column(db.String(255))
    contract_count = db.Column(db.Integer)
    order_count = db.Column(db.Integer)
    total_quantity = db.Column(db.Float)
    total_amount = db.Column(db.Float)
    min_unit_price = db.Column(db.Float)
    max_unit_price = db.Column(db.Float)


class CacheStore(db.Model):
    __tablename__ = "cache_store"
    key = db.Column(db.String(80), primary_key=True)
    value_json = db.Column(db.Text)
    updated_at = db.Column(db.String(30))


class AuthSession(db.Model):
    __tablename__ = "auth_session"
    token = db.Column(db.String(128), primary_key=True)
    expires_at = db.Column(db.Float, index=True)


def clean_text(value):
    if pd.isna(value):
        return None
    text = str(value).strip()
    return text or None


def clean_number(value):
    if pd.isna(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def normalize_header_name(value):
    if value is None:
        return ""
    text = str(value).replace("\u3000", " ").strip()
    return "".join(text.split())


def validate_order_headers(df: pd.DataFrame):
    actual_headers = [normalize_header_name(col) for col in df.columns[: len(ORDER_UPLOAD_COLUMNS)]]
    expected_headers = [normalize_header_name(col) for col in ORDER_UPLOAD_COLUMNS]
    return actual_headers == expected_headers, actual_headers


def safe_json_value(value):
    if isinstance(value, dict):
        return {key: safe_json_value(item) for key, item in value.items()}
    if isinstance(value, list):
        return [safe_json_value(item) for item in value]
    if isinstance(value, tuple):
        return tuple(safe_json_value(item) for item in value)

    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None

    return value


def parse_db_conditions(raw_value):
    if not raw_value:
        return []

    try:
        items = json.loads(raw_value)
    except (TypeError, ValueError, json.JSONDecodeError):
        return []

    if not isinstance(items, list):
        return []

    conditions = []
    for item in items:
        if not isinstance(item, dict):
            continue
        field = clean_text(item.get("field"))
        raw_value = item.get("value")
        values = []
        if isinstance(raw_value, list):
            for single in raw_value:
                normalized = clean_text(single)
                if normalized:
                    values.append(normalized)
        else:
            normalized = clean_text(raw_value)
            if normalized:
                values.append(normalized)

        if field and values:
            conditions.append({"field": field, "value": values})
    return conditions


def parse_db_searches(raw_value):
    if not raw_value:
        return []

    try:
        items = json.loads(raw_value)
    except (TypeError, ValueError, json.JSONDecodeError):
        return []

    if not isinstance(items, list):
        return []

    values = []
    for item in items:
        text = clean_text(item)
        if text:
            values.append(text)
    return values


def apply_db_conditions(query, conditions, field_map):
    clauses = []
    for item in conditions:
        column = field_map.get(item.get("field"))
        values = item.get("value")
        if column is None or not values:
            continue
        if not isinstance(values, list):
            values = [values]
        value_clauses = [column.ilike(f"%{value}%") for value in values if value]
        if value_clauses:
            clauses.append(or_(*value_clauses))

    if clauses:
        query = query.filter(and_(*clauses))
    return query


def build_generic_search_clause(search_value, columns):
    if not search_value:
        return None

    clauses = [column.ilike(f"%{search_value}%") for column in columns]
    return or_(*clauses) if clauses else None


def normalize_part_code(value):
    text = clean_text(value)
    if not text:
        return None
    return re.sub(r"\.0+$", "", text)


def normalize_customer_code(value):
    text = clean_text(value)
    if not text:
        return None
    # Strip trailing hyphens so "CLA-" -> "CLA", "REC-M" stays "REC-M", "CL-" -> "CL"
    normalized = text.rstrip("-").strip()
    return normalized if normalized else text


def customer_filter_clause(value):
    normalized = normalize_customer_code(value)
    if not normalized:
        return None
    return Order.customer_code.ilike(f"%{normalized}%")


def expand_part_lookup_keys(values):
    keys = set()
    for value in values:
        raw = clean_text(value)
        normalized = normalize_part_code(value)

        if raw:
            keys.add(raw)
        if normalized:
            keys.add(normalized)
            if re.fullmatch(r"\d+", normalized):
                keys.add(f"{normalized}.0")
    return keys


def serialize_order(order: Order):
    return {
        "id": order.id,
        "contractNo": order.contract_no,
        "customerCode": normalize_customer_code(order.customer_code),
        "sequence": order.sequence,
        "partNo": normalize_part_code(order.part_no),
        "interchangePartNo": normalize_part_code(order.interchange_part_no),
        "partName": order.part_name,
        "partNameCn": order.part_name_cn,
        "quantity": order.quantity,
        "unitPrice": order.unit_price,
        "totalPrice": order.total_price,
    }


def parse_contract_order_info(contract_no):
    if not contract_no:
        return {"orderYear": None, "orderDate": None}

    match = re.search(r"(\d{6})", str(contract_no))
    if not match:
        return {"orderYear": None, "orderDate": None}

    digits = match.group(1)
    return {
        "orderYear": f"20{digits[:2]}",
        "orderDate": digits[2:],
    }


def get_contract_month(contract_no):
    info = parse_contract_order_info(contract_no)
    if not info["orderDate"]:
        return None

    try:
        return int(str(info["orderDate"])[0:2])
    except (TypeError, ValueError):
        return None


def get_contract_year(contract_no):
    info = parse_contract_order_info(contract_no)
    return info.get("orderYear")


def get_contract_sort_key(contract_no):
    info = parse_contract_order_info(contract_no)
    year = info.get("orderYear")
    order_date = info.get("orderDate")
    if not year or not order_date:
        return None
    return f"{year}{order_date}"


def format_contract_full_date(contract_no):
    info = parse_contract_order_info(contract_no)
    year = info.get("orderYear")
    order_date = info.get("orderDate")
    if not year or not order_date or len(str(order_date)) != 4:
        return None
    return f"{year}-{str(order_date)[:2]}-{str(order_date)[2:]}"


def base_order_query():
    return Order.query.filter(Order.contract_no.isnot(None))


def format_order_date_text(value):
    if not value:
        return ""
    text = str(value)
    return f"{text[:2]}-{text[2:]}" if len(text) == 4 else text


def list_arrival_files():
    candidates = []
    search_dir = ARRIVAL_DIR if ARRIVAL_DIR.exists() else BASE_DIR

    for path in search_dir.glob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        if search_dir == BASE_DIR and path.name == SOURCE_FILE.name:
            continue
        if path.stem.endswith("_imported"):
            continue

        match = re.search(r"(\d{6})", path.name)
        date_key = match.group(1) if match else ""
        candidates.append((date_key, path.name))

    candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
    return [name for _, name in candidates]


def extract_arrival_date_from_filename(filename):
    match = re.search(r"(\d{6})", filename or "")
    return match.group(1) if match else None


def format_adv_date_text(value):
    text = clean_text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d{8}", text):
        return text[2:]
    if re.fullmatch(r"\d{6}", text):
        return text
    return re.sub(r"\D", "", text)[:6]


def build_adv_workbook_bytes(rows, adv_no):
    template_path = BASE_DIR / "templates.xlsx"
    workbook = load_workbook(template_path)
    sheet = workbook[workbook.sheetnames[0]]

    # Delete all rows after row 2 (keep only header rows)
    if sheet.max_row > 2:
        sheet.delete_rows(3, sheet.max_row - 2)
    sheet.merged_cells.ranges.clear()
    thin_side = Side(style="thin")
    thick_side = Side(style="medium")

    # Group by order(contractNo) while preserving original order
    grouped_rows = []
    order_index_map = {}
    for row in rows:
        contract_no = clean_text(row.get("contractNo")) or ""
        if contract_no not in order_index_map:
            order_index_map[contract_no] = len(grouped_rows)
            grouped_rows.append({"contractNo": contract_no, "rows": []})
        grouped_rows[order_index_map[contract_no]]["rows"].append(row)

    if not grouped_rows:
        grouped_rows = [{"contractNo": "", "rows": []}]

    def copy_row_style(source_row, target_row):
        """Copy formatting from source row to target row"""
        source_dim = sheet.row_dimensions[source_row]
        target_dim = sheet.row_dimensions[target_row]
        target_dim.height = source_dim.height

        for col in range(1, 8):
            source_cell = sheet.cell(source_row, col)
            target_cell = sheet.cell(target_row, col)
            target_cell._style = copy(source_cell._style)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.fill = copy(source_cell.fill)
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)

    def set_cell_border(cell, left=None, right=None, top=None, bottom=None):
        current = cell.border
        cell.border = Border(
            left=left if left is not None else current.left,
            right=right if right is not None else current.right,
            top=top if top is not None else current.top,
            bottom=bottom if bottom is not None else current.bottom,
            diagonal=current.diagonal,
            diagonal_direction=current.diagonal_direction,
            vertical=current.vertical,
            horizontal=current.horizontal,
            diagonalUp=current.diagonalUp,
            diagonalDown=current.diagonalDown,
            outline=current.outline,
            start=current.start,
            end=current.end,
        )

    def apply_all_borders(min_row, max_row, min_col=1, max_col=7):
        for row_no in range(min_row, max_row + 1):
            for col_no in range(min_col, max_col + 1):
                cell = sheet.cell(row_no, col_no)
                set_cell_border(cell, left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def apply_thick_outside(min_row, max_row, min_col=1, max_col=7):
        if max_row < min_row:
            return
        for row_no in range(min_row, max_row + 1):
            for col_no in range(min_col, max_col + 1):
                left = thick_side if col_no == min_col else None
                right = thick_side if col_no == max_col else None
                top = thick_side if row_no == min_row else None
                bottom = thick_side if row_no == max_row else None
                if left or right or top or bottom:
                    set_cell_border(sheet.cell(row_no, col_no), left=left, right=right, top=top, bottom=bottom)

    def format_part_code(value):
        text = clean_text(value) or ""
        if re.fullmatch(r"\d{9}", text):
            return f"0{text}"
        return text

    def merge_rows_by_sequence(order_rows):
        merged_rows = []
        row_index_by_key = {}

        for index, row in enumerate(order_rows):
            sequence_key = clean_text(row.get("sequence"))
            if not sequence_key:
                sequence_key = f"__empty_seq__{index}"

            quantity_value = clean_number(row.get("quantity")) or 0
            total_price_value = clean_number(row.get("totalPrice")) or 0

            if sequence_key not in row_index_by_key:
                merged_row = dict(row)
                merged_row["quantity"] = quantity_value
                merged_row["totalPrice"] = total_price_value
                merged_rows.append(merged_row)
                row_index_by_key[sequence_key] = len(merged_rows) - 1
                continue

            target_row = merged_rows[row_index_by_key[sequence_key]]
            target_row["quantity"] = (clean_number(target_row.get("quantity")) or 0) + quantity_value
            target_row["totalPrice"] = (clean_number(target_row.get("totalPrice")) or 0) + total_price_value

        return merged_rows

    part_block_ranges = []
    order_block_ranges = []

    def write_order_block(block_start_row, contract_no, order_rows):
        """Write a single order block starting at block_start_row"""
        resolved_contract_no = clean_text(contract_no)
        if not resolved_contract_no and order_rows:
            resolved_contract_no = clean_text(order_rows[0].get("contractNo"))

        merged_order_rows = merge_rows_by_sequence(order_rows)

        # Set contract number in B1 of this block
        sheet.cell(block_start_row, 2, resolved_contract_no or "")

        # Data starts at row block_start_row + 2 (after ORDER NO and ITEM NO rows)
        data_start_row = block_start_row + 2
        total_row = data_start_row + len(merged_order_rows)

        # Insert rows for data if needed
        if len(merged_order_rows) > 1:
            sheet.insert_rows(data_start_row + 1, amount=len(merged_order_rows) - 1)
            total_row = data_start_row + len(merged_order_rows)

        # Fill in data rows
        for idx, row in enumerate(merged_order_rows):
            row_no = data_start_row + idx
            
            # Copy style from template row 3
            copy_row_style(3, row_no)

            sequence_num = clean_number(row.get("sequence"))
            unit_price = clean_number(row.get("unitPrice"))
            quantity = clean_number(row.get("quantity"))
            total_price = clean_number(row.get("totalPrice"))

            sequence_value = int(round(sequence_num)) if sequence_num is not None else (clean_text(row.get("sequence")) or "")
            unit_price_int = 0 if unit_price is None else int(round(unit_price))
            quantity_int = 0 if quantity is None else int(round(quantity))
            total_price_int = 0 if total_price is None else int(round(total_price))

            sheet.cell(row_no, 1, sequence_value)
            sheet.cell(row_no, 2, format_part_code(row.get("partNo")))
            sheet.cell(row_no, 3, format_part_code(row.get("interchangePartNo")))
            sheet.cell(row_no, 4, row.get("partName") or "")
            sheet.cell(row_no, 5, unit_price_int)
            sheet.cell(row_no, 6, quantity_int)
            sheet.cell(row_no, 7, total_price_int)

            # Set alignment: center for cols 1,2; left for col 4; right for cols 5,6,7
            sheet.cell(row_no, 1).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row_no, 2).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row_no, 3).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row_no, 4).alignment = Alignment(horizontal="left", vertical="center")
            sheet.cell(row_no, 5).alignment = Alignment(horizontal="right", vertical="center")
            sheet.cell(row_no, 6).alignment = Alignment(horizontal="right", vertical="center")
            sheet.cell(row_no, 7).alignment = Alignment(horizontal="right", vertical="center")

            # Set number format for price columns
            sheet.cell(row_no, 5).number_format = "#,##0"
            sheet.cell(row_no, 7).number_format = "#,##0"

        # Create TOTAL row
        sheet.insert_rows(total_row, 1)
        
        # Merge A:E for TOTAL row
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
        
        # Set TOTAL cell formatting: 8pt, bold, italic, underline
        total_cell = sheet.cell(total_row, 1)
        total_cell.value = "TOTAL"
        total_cell.font = Font(size=14, bold=True, italic=True, underline="single")
        total_cell.alignment = Alignment(horizontal="right", vertical="center")

        # Set sum formulas
        sheet.cell(total_row, 6, f"=SUM(F{data_start_row}:F{total_row - 1})")
        sheet.cell(total_row, 7, f"=SUM(G{data_start_row}:G{total_row - 1})")
        
        # Format sum cells: 8pt, bold, italic, underline
        sheet.cell(total_row, 6).font = Font(size=14, bold=True, italic=True, underline="single")
        sheet.cell(total_row, 7).font = Font(size=14, bold=True, italic=True, underline="single")
        sheet.cell(total_row, 6).alignment = Alignment(horizontal="right", vertical="center")
        sheet.cell(total_row, 7).alignment = Alignment(horizontal="right", vertical="center")
        sheet.cell(total_row, 6).number_format = "#,##0"
        sheet.cell(total_row, 7).number_format = "#,##0"

        if total_row - 1 >= data_start_row:
            part_block_ranges.append((data_start_row, total_row - 1, 1, 7))
        order_block_ranges.append((block_start_row + 1, total_row, 1, 7))

        return total_row

    # Write first order block starting at row 1
    current_block_end = write_order_block(
        block_start_row=1,
        contract_no=grouped_rows[0]["contractNo"],
        order_rows=grouped_rows[0]["rows"],
    )

    # Write remaining order blocks with blank rows between
    for group in grouped_rows[1:]:
        next_block_start = current_block_end + 2  # Leave one blank row
        
        # Insert rows for the new block (1 for ORDER NO, 1 for headers, then data rows + 1 for TOTAL)
        num_rows_needed = 2 + len(group["rows"]) + 1
        sheet.insert_rows(next_block_start, num_rows_needed)
        
        # Copy header rows from template (rows 1-2)
        for row_offset in range(2):
            for col in range(1, 8):
                source_cell = sheet.cell(row_offset + 1, col)
                target_cell = sheet.cell(next_block_start + row_offset, col)
                target_cell.value = source_cell.value
                target_cell._style = copy(source_cell._style)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.fill = copy(source_cell.fill)
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
        
        current_block_end = write_order_block(
            block_start_row=next_block_start,
            contract_no=group.get("contractNo") or "",
            order_rows=group.get("rows") or [],
        )

    apply_all_borders(1, current_block_end, 1, 7)
    for min_row, max_row, min_col, max_col in part_block_ranges:
        apply_thick_outside(min_row, max_row, min_col, max_col)
    for min_row, max_row, min_col, max_col in order_block_ranges:
        apply_thick_outside(min_row, max_row, min_col, max_col)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _arrival_row_to_dict(row):
    return {
        "id": row.id,
        "arrivalDate": format_arrival_date_text(row.arrival_date),
        "sourceFile": row.source_file,
        "contractNo": row.contract_no,
        "customerCode": normalize_customer_code(row.customer_code),
        "sequence": clean_text(row.sequence),
        "partNo": normalize_part_code(row.part_no),
        "interchangePartNo": normalize_part_code(row.interchange_part_no),
        "partName": clean_text(row.part_name),
        "quantity": clean_number(row.quantity),
        "unitPrice": clean_number(row.unit_price),
        "totalPrice": clean_number(row.total_price),
    }


def sync_arrivals_from_files():
    files = list_arrival_files()

    existing_sources = {
        value
        for (value,) in db.session.query(ArrivalOrder.source_file)
        .filter(ArrivalOrder.source_file.isnot(None))
        .distinct()
        .all()
        if value
    }

    records = []
    imported_files = 0
    skipped_files = 0

    for file_name in files:
        path = ARRIVAL_DIR / file_name
        if not path.exists():
            path = BASE_DIR / file_name
        if not path.exists():
            continue

        legacy_imported_name = f"{Path(file_name).stem}_imported{Path(file_name).suffix}"
        already_imported = file_name in existing_sources or legacy_imported_name in existing_sources
        if already_imported:
            skipped_files += 1
            continue

        arrival_date = extract_arrival_date_from_filename(file_name)
        frame = pd.read_excel(path, header=None).iloc[:, :8].copy()
        frame.columns = ARRIVAL_COLUMNS
        frame["客户代码"] = frame["合同号"].astype("string").str.split("2", n=1).str[0]

        for _, row in frame.iterrows():
            contract_no = clean_text(row.get("合同号"))
            customer_code = clean_text(row.get("客户代码"))
            sequence = clean_text(row.get("序号"))
            part_no = normalize_part_code(row.get("零件号"))

            if not any([customer_code, contract_no, sequence, part_no]):
                continue

            records.append(
                ArrivalOrder(
                    source_file=file_name,
                    arrival_date=arrival_date,
                    contract_no=contract_no,
                    customer_code=customer_code,
                    sequence=sequence,
                    part_no=part_no,
                    interchange_part_no=normalize_part_code(row.get("互换零件号")),
                    part_name=clean_text(row.get("零件名")),
                    quantity=clean_number(row.get("个数")),
                    unit_price=clean_number(row.get("单价")),
                    total_price=clean_number(row.get("总价")),
                )
            )
        imported_files += 1

    if records:
        db.session.bulk_save_objects(records)
    db.session.commit()

    ARRIVAL_MATCH_CACHE["signature"] = None
    ARRIVAL_MATCH_CACHE["index"] = defaultdict(list)

    return {
        "imported": len(records),
        "files": len(files),
        "importedFiles": imported_files,
        "renamedFiles": 0,
        "skippedFiles": skipped_files,
        "message": "Arrival data synced successfully.",
    }


def _arrival_db_signature():
    total = db.session.query(func.count(ArrivalOrder.id)).scalar() or 0
    max_id = db.session.query(func.max(ArrivalOrder.id)).scalar() or 0
    return (int(total), int(max_id))


def _to_upper_text(value):
    if pd.isna(value):
        return None
    return str(value).strip().upper()


def _normalize_numeric(value):
    if pd.isna(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _arrival_sort_key(value):
    text = clean_text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d{6}", text):
        return f"20{text}"
    if re.fullmatch(r"\d{8}", text):
        return text
    return text


def format_arrival_date_text(value):
    text = clean_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d{6}", text):
        return f"20{text[:2]}-{text[2:4]}-{text[4:]}"
    if re.fullmatch(r"\d{8}", text):
        return f"{text[:4]}-{text[4:6]}-{text[6:]}"
    return text


def get_arrival_match_index():
    signature = _arrival_db_signature()
    if ARRIVAL_MATCH_CACHE["signature"] == signature:
        return ARRIVAL_MATCH_CACHE["index"]

    index = defaultdict(list)

    rows = ArrivalOrder.query.all()
    for row in rows:
        customer_norm = _to_upper_text(normalize_customer_code(row.customer_code))
        contract_norm = _to_upper_text(row.contract_no)
        if not customer_norm or not contract_norm:
            continue

        part_keys = expand_part_lookup_keys([row.part_no, row.interchange_part_no])
        if not part_keys:
            continue

        index[(customer_norm, contract_norm)].append(
            {
                "partKeySet": part_keys,
                "sequence": _normalize_numeric(row.sequence),
                "partNameNorm": _to_upper_text(row.part_name),
                "unitPrice": _normalize_numeric(row.unit_price),
                "quantity": _normalize_numeric(row.quantity) or 0.0,
                "arrivalDate": row.arrival_date,
            }
        )

    ARRIVAL_MATCH_CACHE["signature"] = signature
    ARRIVAL_MATCH_CACHE["index"] = index
    return index


def _match_arrivals_for_order(order, candidates):
    if not candidates:
        return 0.0, []

    order_part_keys = expand_part_lookup_keys([order.part_no, order.interchange_part_no])
    if not order_part_keys:
        return 0.0, []

    filtered = [item for item in candidates if item["partKeySet"] & order_part_keys]
    if not filtered:
        return 0.0, []

    order_sequence = _normalize_numeric(order.sequence)
    if order_sequence is not None:
        narrowed = [item for item in filtered if item.get("sequence") == order_sequence]
        if narrowed:
            filtered = narrowed

    part_name_norm = _to_upper_text(order.part_name)
    if part_name_norm:
        narrowed = [item for item in filtered if item.get("partNameNorm") == part_name_norm]
        if narrowed:
            filtered = narrowed

    unit_price = _normalize_numeric(order.unit_price)
    if unit_price is not None:
        narrowed = [item for item in filtered if item.get("unitPrice") == unit_price]
        if narrowed:
            filtered = narrowed

    total_quantity = sum(float(item.get("quantity") or 0) for item in filtered)
    date_values = sorted(
        {item.get("arrivalDate") for item in filtered if item.get("arrivalDate")},
        key=_arrival_sort_key,
    )
    formatted_dates = [format_arrival_date_text(value) for value in date_values if format_arrival_date_text(value)]
    return round(total_quantity, 2), formatted_dates


def _get_arrival_details_for_order(order, candidates):
    """Get detailed arrival information (time + quantity) for a single order."""
    if not candidates:
        return []

    order_part_keys = expand_part_lookup_keys([order.part_no, order.interchange_part_no])
    if not order_part_keys:
        return []

    filtered = [item for item in candidates if item["partKeySet"] & order_part_keys]
    if not filtered:
        return []

    order_sequence = _normalize_numeric(order.sequence)
    if order_sequence is not None:
        narrowed = [item for item in filtered if item.get("sequence") == order_sequence]
        if narrowed:
            filtered = narrowed

    part_name_norm = _to_upper_text(order.part_name)
    if part_name_norm:
        narrowed = [item for item in filtered if item.get("partNameNorm") == part_name_norm]
        if narrowed:
            filtered = narrowed

    unit_price = _normalize_numeric(order.unit_price)
    if unit_price is not None:
        narrowed = [item for item in filtered if item.get("unitPrice") == unit_price]
        if narrowed:
            filtered = narrowed

    # Group by arrival date and sum quantities
    details_by_date = defaultdict(float)
    for item in filtered:
        arrival_date = item.get("arrivalDate")
        if arrival_date:
            details_by_date[arrival_date] += float(item.get("quantity") or 0)

    # Sort by date and format
    result = []
    for date_str in sorted(details_by_date.keys(), key=_arrival_sort_key):
        formatted_date = format_arrival_date_text(date_str)
        if formatted_date:
            result.append({
                "date": formatted_date,
                "quantity": round(details_by_date[date_str], 2)
            })
    
    return result


def _prepare_history_df():
    order_rows = Order.query.filter(Order.contract_no.isnot(None)).all()
    data = [
        {
            "合同号": row.contract_no,
            "客户代码": normalize_customer_code(row.customer_code) or "",
            "序号": row.sequence,
            "零件号": row.part_no,
            "互换零件号": row.interchange_part_no,
            "零件名": row.part_name,
            "个数": row.quantity,
            "单价": row.unit_price,
            "总价": row.total_price,
        }
        for row in order_rows
    ]
    if not data:
        return pd.DataFrame(columns=["合同号", "客户代码", "序号", "零件号", "互换零件号", "零件名", "个数", "单价", "总价", "零件名_norm", "客户代码_norm", "合同号_norm"])

    data_df = pd.DataFrame(data)

    for col in ["序号", "零件号", "互换零件号", "个数", "单价", "总价"]:
        if col in data_df.columns:
            data_df[col] = pd.to_numeric(data_df[col], errors="coerce")

    data_df["零件名_norm"] = data_df["零件名"].astype("string").str.strip().str.upper()
    data_df["客户代码_norm"] = data_df["客户代码"].astype("string").str.strip().str.upper()
    data_df["合同号_norm"] = data_df["合同号"].astype("string").str.strip().str.upper()
    return data_df


def _get_part_keys(row_like):
    keys = set()
    for col in ["零件号", "互换零件号"]:
        if col not in row_like.index:
            continue
        value = row_like.get(col)
        if pd.notna(value):
            keys.add(value)
    return keys


def _find_part_candidates(source_row, hist_df):
    part_keys = _get_part_keys(source_row)
    if not part_keys:
        return hist_df.iloc[0:0].copy()

    if "互换零件号" in hist_df.columns:
        return hist_df[
            hist_df["零件号"].isin(part_keys)
            | hist_df["互换零件号"].isin(part_keys)
        ].copy()

    return hist_df[hist_df["零件号"].isin(part_keys)].copy()


def analyze_arrivals():
    rows = ArrivalOrder.query.order_by(ArrivalOrder.arrival_date.desc(), ArrivalOrder.id.asc()).all()

    if not rows:
        return {
            "summary": {
                "totalRows": 0,
                "checkedRows": 0,
                "correctRows": 0,
                "errorRows": 0,
                "errorRate": 0,
                "totalFiles": 0,
            },
            "errorFieldStats": [],
            "fileStats": [],
            "customerStats": [],
            "checks": [],
            "errors": [],
        }

    arrival_df = pd.DataFrame(
        [
            {
                "合同号": row.contract_no,
                "序号": row.sequence,
                "零件号": row.part_no,
                "互换零件号": row.interchange_part_no,
                "零件名": row.part_name,
                "个数": row.quantity,
                "单价": row.unit_price,
                "总价": row.total_price,
                "到货日期": row.arrival_date,
                "来源文件": row.source_file,
                "客户代码": row.customer_code,
            }
            for row in rows
        ]
    )

    def _extract_arrival_year(*values):
        for value in values:
            text = clean_text(value)
            if not text:
                continue
            match = re.search(r"(\d{6})", text)
            if match:
                return f"20{match.group(1)[:2]}"
        return None

    arrival_years = {
        "2025": {"files": set(), "rows": 0},
        "2026": {"files": set(), "rows": 0},
    }
    for _, row in arrival_df.iterrows():
        year = _extract_arrival_year(row.get("到货日期"), row.get("来源文件"))
        if year not in arrival_years:
            continue
        arrival_years[year]["rows"] += 1
        source_file = clean_text(row.get("来源文件"))
        if source_file:
            arrival_years[year]["files"].add(source_file)

    history_df = _prepare_history_df()

    df_check = arrival_df[
        arrival_df["合同号"].notna()
        & arrival_df["客户代码"].notna()
        & arrival_df["零件号"].notna()
    ].copy()

    for col in ["序号", "零件号", "互换零件号", "个数", "单价", "总价"]:
        if col in df_check.columns:
            df_check[col] = pd.to_numeric(df_check[col], errors="coerce")

    df_check["零件名_norm"] = df_check["零件名"].astype("string").str.strip().str.upper()
    df_check["客户代码_norm"] = df_check["客户代码"].apply(
        lambda x: (normalize_customer_code(x) or "").upper() if pd.notna(x) else ""
    ).astype("string")
    df_check["合同号_norm"] = df_check["合同号"].astype("string").str.strip().str.upper()

    history_grouped = {
        key: value.copy()
        for key, value in history_df.groupby("客户代码_norm", dropna=False)
    }

    fields_to_check = ["序号", "零件名_norm", "单价"]
    field_name_map = {"零件名_norm": "零件名"}
    check_rows = []

    for idx, row in df_check.iterrows():
        customer = row["客户代码_norm"]
        contract_no = row["合同号_norm"]

        cust_df = history_grouped.get(customer, pd.DataFrame(columns=history_df.columns))
        if cust_df.empty:
            check_rows.append({"index": idx, "检查结果": "有错误", "错误字段": "历史数据中未找到该客户"})
            continue

        contract_df = cust_df[cust_df["合同号_norm"] == contract_no]
        if contract_df.empty:
            check_rows.append({"index": idx, "检查结果": "有错误", "错误字段": "该客户下未找到该合同号"})
            continue

        candidates = _find_part_candidates(row, contract_df)
        if candidates.empty:
            check_rows.append({
                "index": idx,
                "检查结果": "有错误",
                "错误字段": "该合同下未找到该零件号/互换零件号（到货零件未出现在历史合同中）",
            })
            continue

        wrong_fields = []
        for field in fields_to_check:
            if field not in row.index or field not in candidates.columns:
                continue

            value = row[field]
            if field == "零件名_norm":
                ref_values = set(candidates[field].dropna().astype("string").str.strip().str.upper().tolist())
                normalized = None if pd.isna(value) else str(value).strip().upper()
                if normalized is None:
                    if ref_values:
                        wrong_fields.append(field_name_map.get(field, field))
                elif normalized not in ref_values:
                    wrong_fields.append(field_name_map.get(field, field))
            else:
                ref_values = set(candidates[field].dropna().tolist())
                if pd.isna(value):
                    if ref_values:
                        wrong_fields.append(field_name_map.get(field, field))
                elif value not in ref_values:
                    wrong_fields.append(field_name_map.get(field, field))

        check_rows.append({
            "index": idx,
            "检查结果": "全部正确" if not wrong_fields else "有错误",
            "错误字段": "" if not wrong_fields else "、".join(wrong_fields),
        })

    result_df = df_check.join(pd.DataFrame(check_rows).set_index("index")) if check_rows else df_check.copy()

    def _fmt_records(frame, max_rows=3):
        if frame.empty:
            return ""
        cols = ["合同号", "客户代码", "零件号", "互换零件号", "序号", "零件名", "个数", "单价", "总价"]
        use_cols = [col for col in cols if col in frame.columns]
        rows = frame[use_cols].head(max_rows).to_dict("records")
        return " | ".join(str(item) for item in rows)

    def _find_actual_in_data(row):
        customer = row.get("客户代码_norm") or (normalize_customer_code(row.get("客户代码")) or "").upper()
        contract = row.get("合同号_norm") or _to_upper_text(row.get("合同号"))

        cust_df = history_df[history_df["客户代码_norm"] == customer]
        if cust_df.empty:
            return "历史数据中未找到该客户"

        contract_df = cust_df[cust_df["合同号_norm"] == contract]
        if contract_df.empty:
            return "该客户下未找到该合同号"

        matched = _find_part_candidates(row, contract_df)
        if matched.empty:
            return "该合同下未找到该零件号/互换零件号（到货零件未出现在历史合同中）"

        return _fmt_records(matched)

    error_df = result_df[result_df["检查结果"] == "有错误"].copy() if "检查结果" in result_df.columns else result_df.iloc[0:0].copy()
    if not error_df.empty:
        error_df["历史匹配情况"] = error_df.apply(_find_actual_in_data, axis=1)

    error_field_counter = Counter()
    for value in error_df.get("错误字段", pd.Series(dtype="string")).dropna().tolist():
        for field in [item for item in str(value).split("、") if item]:
            error_field_counter[field] += 1

    total_checked = int(len(result_df))
    total_error = int(len(error_df))
    total_correct = max(total_checked - total_error, 0)

    check_items = []
    for idx, row in result_df.head(5000).iterrows():
        check_items.append(
            {
                "rowIndex": int(idx) + 1,
                "arrivalDate": format_arrival_date_text(row.get("到货日期")),
                "sourceFile": row.get("来源文件"),
                "contractNo": row.get("合同号"),
                "customerCode": row.get("客户代码"),
                "partNo": normalize_part_code(row.get("零件号")),
                "interchangePartNo": normalize_part_code(row.get("互换零件号")),
                "sequence": row.get("序号"),
                "partName": row.get("零件名"),
                "quantity": row.get("个数"),
                "unitPrice": row.get("单价"),
                "totalPrice": row.get("总价"),
                "checkResult": row.get("检查结果") or "",
                "errorFields": row.get("错误字段") or "",
                "actualMatch": row.get("历史匹配情况") or "",
            }
        )

    error_items = []
    for idx, row in error_df.head(500).iterrows():
        error_items.append(
            {
                "rowIndex": int(idx) + 1,
                "arrivalDate": format_arrival_date_text(row.get("到货日期")),
                "sourceFile": row.get("来源文件"),
                "contractNo": row.get("合同号"),
                "customerCode": row.get("客户代码"),
                "partNo": normalize_part_code(row.get("零件号")),
                "interchangePartNo": normalize_part_code(row.get("互换零件号")),
                "sequence": row.get("序号"),
                "partName": row.get("零件名"),
                "unitPrice": row.get("单价"),
                "checkResult": row.get("检查结果"),
                "errorFields": row.get("错误字段") or "",
                "actualMatch": row.get("历史匹配情况") or "",
            }
        )

    check_group = {}
    if "来源文件" in result_df.columns and not result_df.empty:
        for source_file, group in result_df.groupby("来源文件", dropna=False):
            key = None if pd.isna(source_file) else str(source_file)
            group_error_rows = int((group["检查结果"] == "有错误").sum()) if "检查结果" in group.columns else 0
            check_group[key] = {
                "checkedRows": int(len(group)),
                "correctRows": int(len(group)) - group_error_rows,
                "errorRows": group_error_rows,
            }

    file_stats = []
    if "来源文件" in arrival_df.columns and not arrival_df.empty:
        for source_file, group in arrival_df.groupby("来源文件", dropna=False):
            key = None if pd.isna(source_file) else str(source_file)
            customer_group = group.iloc[0:0].copy()
            if "客户代码" in group.columns:
                customer_mask = (
                    group["客户代码"].notna()
                    & group["客户代码"].astype("string").str.strip().ne("")
                )
                customer_group = group[customer_mask].copy()

            customer_count = int(customer_group["客户代码"].astype("string").str.strip().nunique()) if "客户代码" in customer_group.columns else 0
            contract_count = int(customer_group["合同号"].dropna().astype("string").str.strip().nunique()) if "合同号" in customer_group.columns else 0

            customer_totals = pd.DataFrame(columns=["客户代码", "个数", "总价"])
            if not customer_group.empty:
                customer_totals = (
                    customer_group.assign(
                        客户代码_key=customer_group["客户代码"].astype("string").str.strip(),
                        个数_num=pd.to_numeric(customer_group.get("个数"), errors="coerce").fillna(0),
                        总价_num=pd.to_numeric(customer_group.get("总价"), errors="coerce").fillna(0),
                    )
                    .groupby("客户代码_key", dropna=False)[["个数_num", "总价_num"]]
                    .sum()
                    .reset_index()
                )

            total_quantity = float(customer_totals["个数_num"].sum()) if "个数_num" in customer_totals.columns else 0.0
            total_amount = float(customer_totals["总价_num"].sum()) if "总价_num" in customer_totals.columns else 0.0
            date_values = sorted(
                {
                    value
                    for value in customer_group.get("到货日期", pd.Series(dtype="string")).dropna().tolist()
                    if clean_text(value)
                },
                key=_arrival_sort_key,
            )
            formatted_dates = [format_arrival_date_text(value) for value in date_values if format_arrival_date_text(value)]

            check_info = check_group.get(key, {"checkedRows": 0, "correctRows": 0, "errorRows": 0})
            file_stats.append(
                {
                    "sourceFile": key,
                    "arrivalDateRange": "、".join(formatted_dates) if formatted_dates else "-",
                    "customerCount": customer_count,
                    "contractCount": contract_count,
                    "totalRows": int(len(customer_group)),
                    "checkedRows": check_info["checkedRows"],
                    "correctRows": check_info["correctRows"],
                    "errorRows": check_info["errorRows"],
                    "totalQuantity": round(total_quantity, 2),
                    "totalAmount": round(total_amount, 2),
                }
            )

    file_stats.sort(key=lambda item: clean_text(item.get("sourceFile")) or "", reverse=True)

    customer_stats = []
    if "客户代码" in arrival_df.columns and not arrival_df.empty:
        customer_frame = arrival_df[
            arrival_df["客户代码"].notna()
            & arrival_df["客户代码"].astype("string").str.strip().ne("")
        ].copy()

        if not customer_frame.empty:
            customer_frame["客户代码_key"] = customer_frame["客户代码"].astype("string").str.strip()
            customer_frame["合同号_key"] = customer_frame["合同号"].astype("string").str.strip()
            customer_frame["来源文件_key"] = customer_frame["来源文件"].astype("string").str.strip()
            customer_frame["到货日期_key"] = customer_frame["到货日期"].astype("string").str.strip()
            customer_frame["个数_num"] = pd.to_numeric(customer_frame.get("个数"), errors="coerce").fillna(0)
            customer_frame["总价_num"] = pd.to_numeric(customer_frame.get("总价"), errors="coerce").fillna(0)

            for customer_code, group in customer_frame.groupby("客户代码_key", dropna=False):
                contract_values = sorted({
                    value for value in group["合同号_key"].tolist() if clean_text(value)
                })
                source_file_values = sorted({
                    value for value in group["来源文件_key"].tolist() if clean_text(value)
                })
                date_values = sorted(
                    {
                        value for value in group["到货日期_key"].tolist() if clean_text(value)
                    },
                    key=_arrival_sort_key,
                )
                formatted_dates = [
                    format_arrival_date_text(value) for value in date_values if format_arrival_date_text(value)
                ]

                customer_stats.append(
                    {
                        "customerCode": normalize_customer_code(customer_code),
                        "arrivalFileCount": int(len(source_file_values)),
                        "arrivalDates": formatted_dates,
                        "contractCount": int(len(contract_values)),
                        "contracts": contract_values,
                        "totalRows": int(len(group)),
                        "totalQuantity": round(float(group["个数_num"].sum()), 2),
                        "totalAmount": round(float(group["总价_num"].sum()), 2),
                    }
                )

    customer_stats.sort(
        key=lambda item: (
            -(float(item.get("totalAmount") or 0)),
            clean_text(item.get("customerCode")) or "",
        )
    )

    return safe_json_value({
        "summary": {
            "totalRows": int(len(arrival_df)),
            "checkedRows": total_checked,
            "correctRows": total_correct,
            "errorRows": total_error,
            "errorRate": round((total_error / total_checked) * 100, 2) if total_checked else 0,
            "totalFiles": int(arrival_df["来源文件"].nunique()) if "来源文件" in arrival_df.columns else 0,
            "arrivalYears": {
                year: {
                    "files": int(len(bucket["files"])),
                    "rows": int(bucket["rows"]),
                }
                for year, bucket in arrival_years.items()
            },
        },
        "errorFieldStats": [
            {"field": key, "count": value}
            for key, value in sorted(error_field_counter.items(), key=lambda item: item[1], reverse=True)
        ],
        "fileStats": file_stats,
        "customerStats": customer_stats,
        "checks": check_items,
        "errors": error_items,
    })


def get_filtered_contract_rows(search="", year=""):
    query = db.session.query(
        Order.contract_no.label("contractNo"),
        Order.customer_code.label("customerCode"),
        func.count(Order.id).label("partItemCount"),
        func.sum(Order.quantity).label("totalQuantity"),
        func.sum(Order.total_price).label("totalAmount"),
    ).filter(Order.contract_no.isnot(None)).group_by(Order.contract_no, Order.customer_code)

    if search:
        query = query.filter(
            or_(
                Order.contract_no.ilike(f"%{search}%"),
                Order.customer_code.ilike(f"%{search}%"),
            )
        )

    items = query.all()
    available_years = sorted(
        {
            info.get("orderYear")
            for info in (parse_contract_order_info(item.contractNo) for item in items)
            if info.get("orderYear")
        },
        reverse=True,
    )

    # Get arrival index for matching
    arrival_index = get_arrival_match_index()

    rows = []
    for item in items:
        order_info = parse_contract_order_info(item.contractNo)
        if year and order_info.get("orderYear") != year:
            continue

        contract_no_norm = _to_upper_text(item.contractNo)
        customer_code_norm = _to_upper_text(normalize_customer_code(item.customerCode))
        
        # Calculate arrival quantities and amounts
        total_arrival_quantity = 0.0
        total_arrival_amount = 0.0
        all_arrival_details = []  # Collect all arrival details for history
        
        # Get all orders for this contract to match against arrivals
        contract_orders = Order.query.filter_by(
            contract_no=item.contractNo,
            customer_code=item.customerCode
        ).all()
        
        arrival_candidates = arrival_index.get((customer_code_norm, contract_no_norm), [])
        
        for order in contract_orders:
            arrival_qty, _ = _match_arrivals_for_order(order, arrival_candidates)
            total_arrival_quantity += arrival_qty
            
            # Calculate matched arrival amount
            if arrival_qty > 0 and order.unit_price:
                # Amount is based on matched quantity and unit price
                total_arrival_amount += arrival_qty * order.unit_price
            
            # Collect arrival details for this order
            details = _get_arrival_details_for_order(order, arrival_candidates)
            all_arrival_details.extend(details)

        total_quantity = float(item.totalQuantity or 0)
        total_amount = round(float(item.totalAmount or 0), 2)
        not_arrival_amount = round(total_amount - total_arrival_amount, 2)
        
        # Calculate arrival ratio
        arrival_ratio = 0.0
        if total_quantity > 0:
            arrival_ratio = round((total_arrival_quantity / total_quantity) * 100, 2)
        
        # Determine arrival status
        if total_arrival_quantity == 0:
            arrival_status = "未到货"
        elif total_arrival_quantity >= total_quantity:
            arrival_status = "到货"
        else:
            arrival_status = "部分到货"
        
        # Deduplicate and sort arrival details by date
        details_dict = defaultdict(float)
        for detail in all_arrival_details:
            details_dict[detail["date"]] += detail["quantity"]
        
        arrival_history = [
            {"date": date, "quantity": round(qty, 2), "ratio": round((qty / total_quantity * 100) if total_quantity > 0 else 0, 2)}
            for date, qty in sorted(details_dict.items())
        ]

        rows.append(
            {
                "contractNo": item.contractNo,
                "customerCode": normalize_customer_code(item.customerCode),
                "partItemCount": item.partItemCount,
                "totalQuantity": round(float(total_quantity), 2),
                "totalAmount": total_amount,
                "arrivalQuantity": round(float(total_arrival_quantity), 2),
                "arrivalAmount": round(float(total_arrival_amount), 2),
                "notArrivalAmount": not_arrival_amount,
                "arrivalRatio": arrival_ratio,
                "arrivalStatus": arrival_status,
                "arrivalHistory": arrival_history,
                **order_info,
            }
        )

    rows.sort(key=lambda row: float(row["totalAmount"] or 0), reverse=True)
    return rows, available_years


def create_period_buckets():
    return {
        "H1": {"contracts": set(), "totalAmount": 0.0, "totalQuantity": 0.0},
        "H2": {"contracts": set(), "totalAmount": 0.0, "totalQuantity": 0.0},
        "Q1": {"contracts": set(), "totalAmount": 0.0, "totalQuantity": 0.0},
        "Q2": {"contracts": set(), "totalAmount": 0.0, "totalQuantity": 0.0},
        "Q3": {"contracts": set(), "totalAmount": 0.0, "totalQuantity": 0.0},
        "Q4": {"contracts": set(), "totalAmount": 0.0, "totalQuantity": 0.0},
    }


def update_period_buckets(buckets, contract_no, month, amount=0, quantity=0):
    if month is None:
        return

    amount = float(amount or 0)
    quantity = float(quantity or 0)

    if 1 <= month <= 6:
        buckets["H1"]["contracts"].add(contract_no)
        buckets["H1"]["totalAmount"] += amount
        buckets["H1"]["totalQuantity"] += quantity
    elif 7 <= month <= 12:
        buckets["H2"]["contracts"].add(contract_no)
        buckets["H2"]["totalAmount"] += amount
        buckets["H2"]["totalQuantity"] += quantity

    if 1 <= month <= 3:
        quarter = "Q1"
    elif 4 <= month <= 6:
        quarter = "Q2"
    elif 7 <= month <= 9:
        quarter = "Q3"
    elif 10 <= month <= 12:
        quarter = "Q4"
    else:
        quarter = None

    if quarter:
        buckets[quarter]["contracts"].add(contract_no)
        buckets[quarter]["totalAmount"] += amount
        buckets[quarter]["totalQuantity"] += quantity


def finalize_period_buckets(buckets):
    return {
        key: {
            "orderCount": len(value["contracts"]),
            "totalAmount": round(value["totalAmount"], 2),
            "totalQuantity": round(value["totalQuantity"], 2),
        }
        for key, value in buckets.items()
    }


def update_price_range(bucket, unit_price):
    if unit_price is None:
        return

    price = float(unit_price)
    bucket["minUnitPrice"] = price if bucket.get("minUnitPrice") is None else min(bucket["minUnitPrice"], price)
    bucket["maxUnitPrice"] = price if bucket.get("maxUnitPrice") is None else max(bucket["maxUnitPrice"], price)


def get_year_bucket(yearly_buckets, year):
    year_key = str(year) if year else "未识别"
    if year_key not in yearly_buckets:
        yearly_buckets[year_key] = {
            "year": year_key,
            "contracts": set(),
            "entryCount": 0,
            "totalAmount": 0.0,
            "totalQuantity": 0.0,
            "partSet": set(),
            "customerSet": set(),
            "partMap": {},
            "customerMap": {},
            "minUnitPrice": None,
            "maxUnitPrice": None,
            "periods": create_period_buckets(),
        }
    return yearly_buckets[year_key]


def serialize_part_summary_map(part_map):
    items = []
    for item in part_map.values():
        items.append(
            {
                "partNo": normalize_part_code(item["partNo"]),
                "interchangePartNo": normalize_part_code(item["interchangePartNo"]),
                "partName": item["partName"],
                "entryCount": item["entryCount"],
                "contractCount": len({x for x in item["contractSet"] if x}),
                "totalQuantity": round(item["totalQuantity"], 2),
                "totalAmount": round(item["totalAmount"], 2),
                "minUnitPrice": None if item["minUnitPrice"] is None else round(item["minUnitPrice"], 2),
                "maxUnitPrice": None if item["maxUnitPrice"] is None else round(item["maxUnitPrice"], 2),
            }
        )
    items.sort(key=lambda x: x["totalAmount"], reverse=True)
    return items


def serialize_customer_summary_map(customer_map):
    items = [
        {
            "customerCode": item["customerCode"],
            "contractCount": len({x for x in item["contractSet"] if x}),
            "entryCount": item["entryCount"],
            "totalQuantity": round(item["totalQuantity"], 2),
            "totalAmount": round(item["totalAmount"], 2),
            "minUnitPrice": None if item["minUnitPrice"] is None else round(item["minUnitPrice"], 2),
            "maxUnitPrice": None if item["maxUnitPrice"] is None else round(item["maxUnitPrice"], 2),
        }
        for item in customer_map.values()
    ]
    items.sort(key=lambda x: x["totalAmount"], reverse=True)
    return items


def finalize_yearly_buckets(yearly_buckets):
    def sort_key(item):
        year_key = item[0]
        try:
            return (0, int(year_key))
        except (TypeError, ValueError):
            return (1, str(year_key))

    results = []
    for year_key, bucket in sorted(yearly_buckets.items(), key=sort_key, reverse=True):
        periods = finalize_period_buckets(bucket["periods"])
        min_price = bucket.get("minUnitPrice")
        max_price = bucket.get("maxUnitPrice")
        price_ratio = None
        if min_price not in (None, 0) and max_price is not None:
            price_ratio = round(max_price / min_price, 4)

        results.append(
            {
                "year": year_key,
                "contractCount": len({x for x in bucket["contracts"] if x}),
                "entryCount": bucket["entryCount"],
                "totalAmount": round(bucket["totalAmount"], 2),
                "totalQuantity": round(bucket["totalQuantity"], 2),
                "partTypeCount": len({x for x in bucket.get("partSet", set()) if x}),
                "customerCount": len({x for x in bucket.get("customerSet", set()) if x}),
                "minUnitPrice": None if min_price is None else round(min_price, 2),
                "maxUnitPrice": None if max_price is None else round(max_price, 2),
                "priceRatio": price_ratio,
                "topParts": serialize_part_summary_map(bucket.get("partMap", {})),
                "customerBreakdown": serialize_customer_summary_map(bucket.get("customerMap", {})),
                "halfyearStats": {
                    "H1": periods["H1"],
                    "H2": periods["H2"],
                },
                "quarterStats": {
                    "Q1": periods["Q1"],
                    "Q2": periods["Q2"],
                    "Q3": periods["Q3"],
                    "Q4": periods["Q4"],
                },
            }
        )
    return results


def sync_orders_from_excel(source_file=SOURCE_FILE):
    if not source_file.exists():
        return {"imported": 0, "source": str(source_file), "message": "Excel source file not found."}

    raw_df = pd.read_excel(source_file)
    is_valid_headers, actual_headers = validate_order_headers(raw_df)
    if not is_valid_headers:
        expected = "\t".join(ORDER_UPLOAD_COLUMNS)
        actual = "\t".join(actual_headers) if actual_headers else "(空)"
        return {
            "imported": 0,
            "source": str(source_file),
            "error": "上传合同信息表失败：第一行表头格式不正确。",
            "message": f"第一行表头必须严格为：{expected}。当前表头为：{actual}",
        }

    rename_map = {
        "合同号": "contract_no",
        "客户代码": "customer_code",
        "序号": "sequence",
        "零件号": "part_no",
        "互换零件号": "interchange_part_no",
        "零件名": "part_name",
        "中文零件名": "part_name_cn",
        "个数": "quantity",
        "单价": "unit_price",
        "总价": "total_price",
    }
    df = raw_df.rename(columns=rename_map).copy()

    if "customer_code" not in df.columns and "contract_no" in df.columns:
        df["customer_code"] = df["contract_no"].astype("string").str.split("2", n=1).str[0]

    db.session.query(Order).delete()

    orders = []
    for _, row in df.iterrows():
        orders.append(
            Order(
                contract_no=clean_text(row.get("contract_no")),
                customer_code=normalize_customer_code(row.get("customer_code")),
                sequence=clean_text(row.get("sequence")),
                part_no=normalize_part_code(row.get("part_no")),
                interchange_part_no=normalize_part_code(row.get("interchange_part_no")),
                part_name=clean_text(row.get("part_name")),
                part_name_cn=clean_text(row.get("part_name_cn")),
                quantity=clean_number(row.get("quantity")),
                unit_price=clean_number(row.get("unit_price")),
                total_price=clean_number(row.get("total_price")),
            )
        )

    if orders:
        db.session.bulk_save_objects(orders)
    db.session.commit()

    return {"imported": len(orders), "source": str(source_file), "message": "Orders synced successfully."}


def rebuild_contract_summary():
    all_orders = Order.query.filter(Order.contract_no.isnot(None)).all()
    contract_groups = defaultdict(list)
    for order in all_orders:
        contract_groups[(order.contract_no, order.customer_code)].append(order)

    arrival_index = get_arrival_match_index()
    records = []
    for (contract_no, customer_code), order_list in contract_groups.items():
        order_info = parse_contract_order_info(contract_no)
        customer_norm = _to_upper_text(normalize_customer_code(customer_code))
        contract_norm = _to_upper_text(contract_no)
        arrival_candidates = arrival_index.get((customer_norm, contract_norm), [])

        total_quantity = sum(float(o.quantity or 0) for o in order_list)
        total_amount = sum(float(o.total_price or 0) for o in order_list)
        total_arrival_quantity = 0.0
        total_arrival_amount = 0.0
        details_dict = defaultdict(float)

        for order in order_list:
            qty, _ = _match_arrivals_for_order(order, arrival_candidates)
            total_arrival_quantity += qty
            if qty > 0 and order.unit_price:
                total_arrival_amount += qty * order.unit_price
            for detail in _get_arrival_details_for_order(order, arrival_candidates):
                details_dict[detail["date"]] += detail["quantity"]

        arrival_ratio = round((total_arrival_quantity / total_quantity) * 100, 2) if total_quantity > 0 else 0.0
        if total_arrival_quantity == 0:
            arrival_status = "未到货"
        elif total_arrival_quantity >= total_quantity:
            arrival_status = "到货"
        else:
            arrival_status = "部分到货"

        arrival_history = [
            {"date": d, "quantity": round(q, 2), "ratio": round((q / total_quantity * 100) if total_quantity > 0 else 0, 2)}
            for d, q in sorted(details_dict.items())
        ]

        records.append(ContractSummary(
            contract_no=contract_no,
            customer_code=normalize_customer_code(customer_code),
            order_year=order_info.get("orderYear"),
            order_date=order_info.get("orderDate"),
            part_item_count=len(order_list),
            total_quantity=round(total_quantity, 2),
            total_amount=round(total_amount, 2),
            arrival_quantity=round(total_arrival_quantity, 2),
            arrival_amount=round(total_arrival_amount, 2),
            not_arrival_amount=round(total_amount - total_arrival_amount, 2),
            arrival_ratio=arrival_ratio,
            arrival_status=arrival_status,
            arrival_history_json=json.dumps(arrival_history),
        ))

    db.session.query(ContractSummary).delete()
    db.session.bulk_save_objects(records)
    db.session.commit()


def rebuild_customer_summary():
    summary_map = {}
    for row in base_order_query().all():
        customer_code = normalize_customer_code(row.customer_code) or "未知"
        item = summary_map.setdefault(customer_code, {
            "contractSet": set(), "orderCount": 0, "totalAmount": 0.0,
            "latestSortKey": None, "latestContractTime": None, "latestContractNo": None,
        })
        item["contractSet"].add(row.contract_no)
        item["orderCount"] += 1
        item["totalAmount"] += float(row.total_price or 0)
        sort_key = get_contract_sort_key(row.contract_no)
        if sort_key and (item["latestSortKey"] is None or sort_key > item["latestSortKey"]):
            item["latestSortKey"] = sort_key
            item["latestContractTime"] = format_contract_full_date(row.contract_no)
            item["latestContractNo"] = row.contract_no

    records = [
        CustomerSummary(
            customer_code=code,
            contract_count=len({x for x in item["contractSet"] if x}),
            order_count=item["orderCount"],
            total_amount=round(item["totalAmount"], 2),
            latest_contract_time=item["latestContractTime"],
            latest_contract_no=item["latestContractNo"],
        )
        for code, item in summary_map.items()
    ]
    db.session.query(CustomerSummary).delete()
    db.session.bulk_save_objects(records)
    db.session.commit()


def rebuild_part_summary():
    part_map = {}
    for row in base_order_query().all():
        key = (
            normalize_part_code(row.part_no) or "",
            normalize_part_code(row.interchange_part_no) or "",
            clean_text(row.part_name) or "",
        )
        item = part_map.setdefault(key, {
            "contractSet": set(), "orderCount": 0,
            "totalQuantity": 0.0, "totalAmount": 0.0,
            "minUnitPrice": None, "maxUnitPrice": None,
        })
        item["contractSet"].add(row.contract_no)
        item["orderCount"] += 1
        item["totalQuantity"] += float(row.quantity or 0)
        item["totalAmount"] += float(row.total_price or 0)
        if row.unit_price is not None:
            p = float(row.unit_price)
            item["minUnitPrice"] = p if item["minUnitPrice"] is None else min(item["minUnitPrice"], p)
            item["maxUnitPrice"] = p if item["maxUnitPrice"] is None else max(item["maxUnitPrice"], p)

    records = [
        PartSummary(
            part_no=key[0] or None,
            interchange_part_no=key[1] or None,
            part_name=key[2] or None,
            contract_count=len({x for x in item["contractSet"] if x}),
            order_count=item["orderCount"],
            total_quantity=round(item["totalQuantity"], 2),
            total_amount=round(item["totalAmount"], 2),
            min_unit_price=None if item["minUnitPrice"] is None else round(item["minUnitPrice"], 2),
            max_unit_price=None if item["maxUnitPrice"] is None else round(item["maxUnitPrice"], 2),
        )
        for key, item in part_map.items()
    ]
    db.session.query(PartSummary).delete()
    db.session.bulk_save_objects(records)
    db.session.commit()


def rebuild_arrival_analysis_cache():
    from datetime import datetime
    result = analyze_arrivals()
    row = db.session.get(CacheStore, "arrival_analysis")
    if row is None:
        row = CacheStore(key="arrival_analysis")
        db.session.add(row)
    row.value_json = json.dumps(result, ensure_ascii=False, default=str)
    row.updated_at = datetime.utcnow().isoformat()
    db.session.commit()


def rebuild_all_summaries():
    rebuild_contract_summary()
    rebuild_customer_summary()
    rebuild_part_summary()
    rebuild_arrival_analysis_cache()


def create_app():
    app = Flask(__name__)
    app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{DB_PATH}"
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

    CORS(app)
    db.init_app(app)
    _setup_auth()

    with app.app_context():
        db.create_all()
        if Order.query.count() == 0:
            sync_orders_from_excel()
        if ArrivalOrder.query.count() == 0:
            sync_arrivals_from_files()
        if ContractSummary.query.count() == 0:
            rebuild_all_summaries()
        elif db.session.get(CacheStore, "arrival_analysis") is None:
            rebuild_arrival_analysis_cache()

    # ── Auth middleware ────────────────────────────────────────────────────────
    _PUBLIC_PATHS = {"/api/health", "/api/login", "/api/logout"}

    @app.before_request
    def check_auth():
        if request.path in _PUBLIC_PATHS or not request.path.startswith("/api/"):
            return None
        raw = request.headers.get("Authorization", "")
        token = raw[7:] if raw.startswith("Bearer ") else raw
        if not token or not _is_valid_session(token):
            return jsonify({"message": "未授权，请重新登录"}), 401

    @app.post("/api/login")
    def login():
        data = request.get_json(silent=True) or {}
        code = str(data.get("code", "")).strip()
        ip = request.remote_addr or "unknown"

        if _is_rate_limited(ip):
            return jsonify({"message": "登录尝试过于频繁，请 15 分钟后再试"}), 429

        if not _verify_code(code):
            _record_failed(ip)
            remaining = 5 - len(_FAILED_ATTEMPTS.get(ip, []))
            return jsonify({"message": f"登录码错误，还有 {max(remaining, 0)} 次机会"}), 401

        token = secrets.token_hex(32)           # 64-char session token
        _clean_sessions()
        db.session.add(AuthSession(token=token, expires_at=time.time() + 86400))
        db.session.commit()
        return jsonify({"token": token})

    @app.post("/api/logout")
    def logout():
        raw = request.headers.get("Authorization", "")
        token = raw[7:] if raw.startswith("Bearer ") else raw
        session_row = db.session.get(AuthSession, token)
        if session_row is not None:
            db.session.delete(session_row)
            db.session.commit()
        return jsonify({"message": "已退出登录"})
    # ── End Auth middleware ────────────────────────────────────────────────────

    @app.get("/api/health")
    def health():
        return jsonify({"status": "ok"})

    def _get_arrival_analysis_result():
        row = db.session.get(CacheStore, "arrival_analysis")
        if row and row.value_json:
            return json.loads(row.value_json)
        return analyze_arrivals()

    @app.get("/api/arrival-analysis")
    def get_arrival_analysis():
        row = db.session.get(CacheStore, "arrival_analysis")
        if row and row.value_json:
            return row.value_json, 200, {"Content-Type": "application/json"}
        return jsonify(analyze_arrivals())

    @app.get("/api/arrival-analysis/export-errors")
    def export_arrival_errors():
        result = _get_arrival_analysis_result()

        errors = result.get("errors", [])
        export_rows = [
            {
                "行号": item.get("rowIndex") or "",
                "到货时间": item.get("arrivalDate") or "",
                "来源文件": item.get("sourceFile") or "",
                "合同号": item.get("contractNo") or "",
                "客户": item.get("customerCode") or "",
                "零件号": item.get("partNo") or "",
                "互换零件号": item.get("interchangePartNo") or "",
                "序号": item.get("sequence") or "",
                "零件名": item.get("partName") or "",
                "单价": item.get("unitPrice") or "",
                "错误字段": item.get("errorFields") or "",
                "历史匹配情况": item.get("actualMatch") or "",
            }
            for item in errors
        ]

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            pd.DataFrame(export_rows).to_excel(writer, sheet_name="arrival_errors", index=False)

        output.seek(0)
        download_name = "到货汇总_错误明细.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.get("/api/arrival-analysis/export-summary")
    def export_arrival_summary():
        result = _get_arrival_analysis_result()

        file_rows = [
            {
                "来源文件": item.get("sourceFile") or "",
                "到货日期": item.get("arrivalDateRange") or "",
                "到货客户数": item.get("customerCount") or 0,
                "到货合同数": item.get("contractCount") or 0,
                "到货行数": item.get("totalRows") or 0,
                "有效检查行数": item.get("checkedRows") or 0,
                "错误行数": item.get("errorRows") or 0,
                "到货总个数": item.get("totalQuantity") or 0,
                "到货总金额": item.get("totalAmount") or 0,
            }
            for item in result.get("fileStats", [])
        ]

        customer_rows = [
            {
                "客户编号": item.get("customerCode") or "",
                "到货文件数": item.get("arrivalFileCount") or 0,
                "到货日期": "、".join(item.get("arrivalDates") or []),
                "到货合同数": item.get("contractCount") or 0,
                "到货涵盖合同": "、".join(item.get("contracts") or []),
                "到货行数": item.get("totalRows") or 0,
                "到货总个数": item.get("totalQuantity") or 0,
                "到货总金额": item.get("totalAmount") or 0,
            }
            for item in result.get("customerStats", [])
        ]

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            pd.DataFrame(file_rows).to_excel(writer, sheet_name="按文件汇总", index=False)
            pd.DataFrame(customer_rows).to_excel(writer, sheet_name="按客户汇总", index=False)

        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name="到货统计汇总.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.get("/api/arrival-customer/export")
    def export_arrival_customer():
        customer_code = request.args.get("customerCode", "").strip()
        if not customer_code:
            return jsonify({"message": "缺少 customerCode 参数"}), 400

        rows = (
            ArrivalOrder.query.filter(
                ArrivalOrder.customer_code.isnot(None),
                func.trim(ArrivalOrder.customer_code) == customer_code,
            )
            .order_by(ArrivalOrder.arrival_date.desc(), ArrivalOrder.contract_no.asc(), ArrivalOrder.id.asc())
            .all()
        )

        items = [_arrival_row_to_dict(row) for row in rows]
        contract_values = sorted({item.get("contractNo") for item in items if item.get("contractNo")})
        date_values = sorted(
            {item.get("arrivalDate") for item in items if item.get("arrivalDate")},
            key=_arrival_sort_key,
        )
        formatted_dates = [
            format_arrival_date_text(value) for value in date_values if format_arrival_date_text(value)
        ]

        summary_rows = [
            {
                "客户编号": customer_code,
                "到货文件数": len({item.get("sourceFile") for item in items if item.get("sourceFile")}),
                "到货日期": "、".join(formatted_dates),
                "到货合同数": len(contract_values),
                "到货涵盖合同": "、".join(contract_values),
                "到货行数": len(items),
                "到货总个数": round(sum(float(item.get("quantity") or 0) for item in items), 2),
                "到货总金额": round(sum(float(item.get("totalPrice") or 0) for item in items), 2),
            }
        ]

        detail_rows = [
            {
                "到货日期": item.get("arrivalDate") or "",
                "来源文件": item.get("sourceFile") or "",
                "合同号": item.get("contractNo") or "",
                "序号": item.get("sequence") or "",
                "零件号": item.get("partNo") or "",
                "互换零件号": item.get("interchangePartNo") or "",
                "零件名": item.get("partName") or "",
                "单价": item.get("unitPrice") or 0,
                "个数": item.get("quantity") or 0,
                "总价": item.get("totalPrice") or 0,
            }
            for item in items
        ]

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            pd.DataFrame(summary_rows).to_excel(writer, sheet_name="客户汇总", index=False)
            pd.DataFrame(detail_rows).to_excel(writer, sheet_name="到货明细", index=False)

        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=f"{customer_code}_到货统计.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.get("/api/arrival-file-detail")
    def get_arrival_file_detail():
        source_file = request.args.get("sourceFile", "").strip()
        if not source_file:
            return jsonify({"message": "缺少 sourceFile 参数", "summary": None, "items": []}), 400

        rows = (
            ArrivalOrder.query.filter(
                ArrivalOrder.source_file == source_file,
                ArrivalOrder.customer_code.isnot(None),
                func.trim(ArrivalOrder.customer_code) != "",
            )
            .order_by(ArrivalOrder.customer_code.asc(), ArrivalOrder.contract_no.asc(), ArrivalOrder.id.asc())
            .all()
        )

        items = [_arrival_row_to_dict(row) for row in rows]
        summary = {
            "sourceFile": source_file,
            "totalRows": len(items),
            "customerCount": len({item.get("customerCode") for item in items if item.get("customerCode")}),
            "contractCount": len({item.get("contractNo") for item in items if item.get("contractNo")}),
            "totalQuantity": round(sum(float(item.get("quantity") or 0) for item in items), 2),
            "totalAmount": round(sum(float(item.get("totalPrice") or 0) for item in items), 2),
        }
        return jsonify(safe_json_value({"summary": summary, "items": items}))

    @app.get("/api/arrival-file/export-adv")
    def export_arrival_file_adv():
        source_file = request.args.get("sourceFile", "").strip()
        if not source_file:
            return jsonify({"message": "缺少 sourceFile 参数"}), 400

        template_path = BASE_DIR / "templates.xlsx"
        if not template_path.exists():
            return jsonify({"message": "templates.xlsx 不存在"}), 400

        # Query arrival rows only to get contract numbers and arrival date;
        # the actual ADV content will come from the Order (contract) table.
        arrival_rows = (
            ArrivalOrder.query.filter(
                ArrivalOrder.source_file == source_file,
                ArrivalOrder.customer_code.isnot(None),
                func.trim(ArrivalOrder.customer_code) != "",
            )
            .order_by(ArrivalOrder.customer_code.asc(), ArrivalOrder.contract_no.asc(), ArrivalOrder.id.asc())
            .all()
        )
        if not arrival_rows:
            return jsonify({"message": "该到货文件无客户数据"}), 404

        first_date = next((clean_text(row.arrival_date) for row in arrival_rows if clean_text(row.arrival_date)), "")
        adv_date = format_adv_date_text(first_date)

        # Collect distinct contract numbers referenced by the arrival file
        contract_nos = list({row.contract_no for row in arrival_rows if row.contract_no})
        if not contract_nos:
            return jsonify({"message": "该到货文件无合同号数据"}), 404

        # Use contract (Order) data as the authoritative source for ADV content
        order_rows = (
            Order.query.filter(
                Order.contract_no.in_(contract_nos),
                Order.customer_code.isnot(None),
                func.trim(Order.customer_code) != "",
            )
            .order_by(Order.customer_code.asc(), Order.contract_no.asc(), Order.id.asc())
            .all()
        )
        if not order_rows:
            return jsonify({"message": "该到货文件对应合同无数据"}), 404

        customer_rows = defaultdict(list)
        for row in order_rows:
            item = {
                "id": row.id,
                "contractNo": row.contract_no,
                "customerCode": normalize_customer_code(row.customer_code),
                "sequence": clean_text(row.sequence),
                "partNo": normalize_part_code(row.part_no),
                "interchangePartNo": normalize_part_code(row.interchange_part_no),
                "partName": clean_text(row.part_name),
                "quantity": clean_number(row.quantity),
                "unitPrice": clean_number(row.unit_price),
                "totalPrice": clean_number(row.total_price),
            }
            customer = item.get("customerCode")
            if not customer:
                continue
            customer_rows[customer].append(item)

        if not customer_rows:
            return jsonify({"message": "该到货文件对应合同无客户数据"}), 404

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zip_file:
            for customer, items in sorted(customer_rows.items(), key=lambda x: x[0]):
                adv_no = f"ADV{adv_date}{customer}" if adv_date else f"ADV{customer}"
                workbook_bytes = build_adv_workbook_bytes(items, adv_no)
                zip_file.writestr(f"{adv_no}.xlsx", workbook_bytes.getvalue())

        zip_buffer.seek(0)
        zip_name = f"ADV_{Path(source_file).stem}.zip"
        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=zip_name,
            mimetype="application/zip",
        )

    @app.get("/api/stats")
    def stats():
        total_orders = db.session.query(func.count(Order.id)).scalar() or 0
        total_contracts = ContractSummary.query.count()
        total_amount = db.session.query(func.sum(ContractSummary.total_amount)).scalar() or 0

        customers = (
            CustomerSummary.query
            .order_by(CustomerSummary.order_count.desc())
            .all()
        )
        total_customers = len(customers)
        top_customers = [{"customerCode": c.customer_code, "count": c.order_count} for c in customers[:5]]

        return jsonify(
            {
                "totalOrders": total_orders,
                "totalCustomers": total_customers,
                "totalContracts": total_contracts,
                "totalAmount": round(float(total_amount), 2),
                "topCustomers": top_customers,
            }
        )

    @app.get("/api/db-check")
    def db_check():
        order_total = db.session.query(func.count(Order.id)).scalar() or 0
        arrival_total = db.session.query(func.count(ArrivalOrder.id)).scalar() or 0

        contract_total = (
            db.session.query(func.count(func.distinct(Order.contract_no)))
            .filter(Order.contract_no.isnot(None), func.trim(Order.contract_no) != "")
            .scalar()
            or 0
        )
        customer_total = (
            db.session.query(func.count(func.distinct(Order.customer_code)))
            .filter(Order.customer_code.isnot(None), func.trim(Order.customer_code) != "")
            .scalar()
            or 0
        )
        arrival_file_total = (
            db.session.query(func.count(func.distinct(ArrivalOrder.source_file)))
            .filter(ArrivalOrder.source_file.isnot(None), func.trim(ArrivalOrder.source_file) != "")
            .scalar()
            or 0
        )
        imported_arrival_rows = (
            db.session.query(func.count(ArrivalOrder.id))
            .filter(ArrivalOrder.source_file.ilike("%_imported%"))
            .scalar()
            or 0
        )

        source_file_stats = [
            {
                "sourceFile": source_file,
                "rows": int(row_count or 0),
            }
            for source_file, row_count in (
                db.session.query(ArrivalOrder.source_file, func.count(ArrivalOrder.id))
                .filter(ArrivalOrder.source_file.isnot(None), func.trim(ArrivalOrder.source_file) != "")
                .group_by(ArrivalOrder.source_file)
                .order_by(func.count(ArrivalOrder.id).desc(), ArrivalOrder.source_file.asc())
                .limit(30)
                .all()
            )
        ]

        latest_orders = [
            {
                "id": row.id,
                "contractNo": row.contract_no,
                "customerCode": normalize_customer_code(row.customer_code),
                "sequence": clean_text(row.sequence),
                "partNo": normalize_part_code(row.part_no),
                "partName": clean_text(row.part_name),
                "quantity": clean_number(row.quantity),
                "unitPrice": clean_number(row.unit_price),
                "totalPrice": clean_number(row.total_price),
            }
            for row in Order.query.order_by(Order.id.desc()).limit(20).all()
        ]

        latest_arrivals = [
            {
                "id": row.id,
                "arrivalDate": format_arrival_date_text(row.arrival_date),
                "sourceFile": row.source_file,
                "contractNo": row.contract_no,
                "customerCode": normalize_customer_code(row.customer_code),
                "sequence": clean_text(row.sequence),
                "partNo": normalize_part_code(row.part_no),
                "partName": clean_text(row.part_name),
                "quantity": clean_number(row.quantity),
                "totalPrice": clean_number(row.total_price),
            }
            for row in ArrivalOrder.query.order_by(ArrivalOrder.id.desc()).limit(20).all()
        ]

        return jsonify(
            safe_json_value(
                {
                    "summary": {
                        "orderRows": int(order_total),
                        "arrivalRows": int(arrival_total),
                        "contractCount": int(contract_total),
                        "customerCount": int(customer_total),
                        "arrivalFileCount": int(arrival_file_total),
                        "importedArrivalRows": int(imported_arrival_rows),
                    },
                    "sourceFileStats": source_file_stats,
                    "latestOrders": latest_orders,
                    "latestArrivals": latest_arrivals,
                }
            )
        )

    @app.get("/api/db-orders")
    def db_orders_list():
        search = request.args.get("search", "").strip()
        refine = request.args.get("refine", "").strip()
        searches = parse_db_searches(request.args.get("searches"))
        conditions = parse_db_conditions(request.args.get("conditions"))
        page = max(int(request.args.get("page", 1)), 1)
        page_size = min(max(int(request.args.get("pageSize", 30)), 1), 100)
        query = Order.query
        order_search_columns = [
            Order.contract_no,
            Order.customer_code,
            Order.sequence,
            Order.part_no,
            Order.interchange_part_no,
            Order.part_name,
            Order.part_name_cn,
        ]
        if searches:
            for value in searches:
                clause = build_generic_search_clause(value, order_search_columns)
                if clause is not None:
                    query = query.filter(clause)
        else:
            search_clause = build_generic_search_clause(search, order_search_columns)
            refine_clause = build_generic_search_clause(refine, order_search_columns)
            if search_clause is not None:
                query = query.filter(search_clause)
            if refine_clause is not None:
                query = query.filter(refine_clause)
        query = apply_db_conditions(
            query,
            conditions,
            {
                "contractNo": Order.contract_no,
                "customerCode": Order.customer_code,
                "sequence": Order.sequence,
                "partNo": Order.part_no,
                "interchangePartNo": Order.interchange_part_no,
                "partName": Order.part_name,
                "partNameCn": Order.part_name_cn,
            },
        )
        total = query.count()
        items = query.order_by(Order.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
        return jsonify({"items": [{"id": r.id, "contractNo": r.contract_no, "customerCode": r.customer_code, "sequence": clean_text(r.sequence), "partNo": normalize_part_code(r.part_no), "interchangePartNo": normalize_part_code(r.interchange_part_no), "partName": clean_text(r.part_name), "partNameCn": clean_text(r.part_name_cn), "quantity": clean_number(r.quantity), "unitPrice": clean_number(r.unit_price), "totalPrice": clean_number(r.total_price)} for r in items], "total": total, "page": page, "pageSize": page_size})

    @app.get("/api/db-filter-options")
    def db_filter_options():
        data_type = request.args.get("type", "").strip().lower()
        field = request.args.get("field", "").strip()

        mapping = {
            "order": {
                "contractNo": Order.contract_no,
                "customerCode": Order.customer_code,
                "sequence": Order.sequence,
                "partNo": Order.part_no,
                "interchangePartNo": Order.interchange_part_no,
                "partName": Order.part_name,
                "partNameCn": Order.part_name_cn,
            },
            "arrival": {
                "sourceFile": ArrivalOrder.source_file,
                "arrivalDate": ArrivalOrder.arrival_date,
                "contractNo": ArrivalOrder.contract_no,
                "customerCode": ArrivalOrder.customer_code,
                "sequence": ArrivalOrder.sequence,
                "partNo": ArrivalOrder.part_no,
                "interchangePartNo": ArrivalOrder.interchange_part_no,
                "partName": ArrivalOrder.part_name,
            },
        }

        field_map = mapping.get(data_type)
        if not field_map or field not in field_map:
            return jsonify({"items": []})

        column = field_map[field]
        rows = db.session.query(column).filter(column.isnot(None)).distinct().order_by(column.asc()).limit(500).all()
        items = [clean_text(row[0]) for row in rows if clean_text(row[0])]
        return jsonify({"items": items})

    @app.post("/api/db-order")
    def db_order_create():
        data = request.get_json() or {}
        order = Order(contract_no=clean_text(data.get("contractNo")), customer_code=clean_text(data.get("customerCode")), sequence=clean_text(data.get("sequence")), part_no=clean_text(data.get("partNo")), interchange_part_no=clean_text(data.get("interchangePartNo")), part_name=clean_text(data.get("partName")), part_name_cn=clean_text(data.get("partNameCn")), quantity=clean_number(data.get("quantity")), unit_price=clean_number(data.get("unitPrice")), total_price=clean_number(data.get("totalPrice")))
        db.session.add(order)
        db.session.commit()
        ARRIVAL_MATCH_CACHE["signature"] = None
        return jsonify({"id": order.id, "message": "创建成功"}), 201

    @app.put("/api/db-order/<int:record_id>")
    def db_order_update(record_id):
        order = db.session.get(Order, record_id)
        if not order:
            return jsonify({"message": "记录不存在"}), 404
        data = request.get_json() or {}
        if "contractNo" in data: order.contract_no = clean_text(data["contractNo"])
        if "customerCode" in data: order.customer_code = clean_text(data["customerCode"])
        if "sequence" in data: order.sequence = clean_text(data["sequence"])
        if "partNo" in data: order.part_no = clean_text(data["partNo"])
        if "interchangePartNo" in data: order.interchange_part_no = clean_text(data["interchangePartNo"])
        if "partName" in data: order.part_name = clean_text(data["partName"])
        if "partNameCn" in data: order.part_name_cn = clean_text(data["partNameCn"])
        if "quantity" in data: order.quantity = clean_number(data["quantity"])
        if "unitPrice" in data: order.unit_price = clean_number(data["unitPrice"])
        if "totalPrice" in data: order.total_price = clean_number(data["totalPrice"])
        db.session.commit()
        ARRIVAL_MATCH_CACHE["signature"] = None
        return jsonify({"message": "更新成功"})

    @app.delete("/api/db-order/<int:record_id>")
    def db_order_delete(record_id):
        order = db.session.get(Order, record_id)
        if not order:
            return jsonify({"message": "记录不存在"}), 404
        db.session.delete(order)
        db.session.commit()
        ARRIVAL_MATCH_CACHE["signature"] = None
        return jsonify({"message": "删除成功"})

    @app.get("/api/db-arrivals")
    def db_arrivals_list():
        search = request.args.get("search", "").strip()
        refine = request.args.get("refine", "").strip()
        searches = parse_db_searches(request.args.get("searches"))
        conditions = parse_db_conditions(request.args.get("conditions"))
        page = max(int(request.args.get("page", 1)), 1)
        page_size = min(max(int(request.args.get("pageSize", 30)), 1), 100)
        query = ArrivalOrder.query
        arrival_search_columns = [
            ArrivalOrder.source_file,
            ArrivalOrder.arrival_date,
            ArrivalOrder.contract_no,
            ArrivalOrder.customer_code,
            ArrivalOrder.sequence,
            ArrivalOrder.part_no,
            ArrivalOrder.interchange_part_no,
            ArrivalOrder.part_name,
        ]
        if searches:
            for value in searches:
                clause = build_generic_search_clause(value, arrival_search_columns)
                if clause is not None:
                    query = query.filter(clause)
        else:
            search_clause = build_generic_search_clause(search, arrival_search_columns)
            refine_clause = build_generic_search_clause(refine, arrival_search_columns)
            if search_clause is not None:
                query = query.filter(search_clause)
            if refine_clause is not None:
                query = query.filter(refine_clause)
        query = apply_db_conditions(
            query,
            conditions,
            {
                "sourceFile": ArrivalOrder.source_file,
                "arrivalDate": ArrivalOrder.arrival_date,
                "contractNo": ArrivalOrder.contract_no,
                "customerCode": ArrivalOrder.customer_code,
                "sequence": ArrivalOrder.sequence,
                "partNo": ArrivalOrder.part_no,
                "interchangePartNo": ArrivalOrder.interchange_part_no,
                "partName": ArrivalOrder.part_name,
            },
        )
        total = query.count()
        items = query.order_by(ArrivalOrder.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
        return jsonify({"items": [{"id": r.id, "sourceFile": r.source_file, "arrivalDate": r.arrival_date, "contractNo": r.contract_no, "customerCode": r.customer_code, "sequence": clean_text(r.sequence), "partNo": normalize_part_code(r.part_no), "interchangePartNo": normalize_part_code(r.interchange_part_no), "partName": clean_text(r.part_name), "quantity": clean_number(r.quantity), "unitPrice": clean_number(r.unit_price), "totalPrice": clean_number(r.total_price)} for r in items], "total": total, "page": page, "pageSize": page_size})

    @app.post("/api/db-arrival")
    def db_arrival_create():
        data = request.get_json() or {}
        arrival = ArrivalOrder(source_file=clean_text(data.get("sourceFile")), arrival_date=clean_text(data.get("arrivalDate")), contract_no=clean_text(data.get("contractNo")), customer_code=clean_text(data.get("customerCode")), sequence=clean_text(data.get("sequence")), part_no=clean_text(data.get("partNo")), interchange_part_no=clean_text(data.get("interchangePartNo")), part_name=clean_text(data.get("partName")), quantity=clean_number(data.get("quantity")), unit_price=clean_number(data.get("unitPrice")), total_price=clean_number(data.get("totalPrice")))
        db.session.add(arrival)
        db.session.commit()
        ARRIVAL_MATCH_CACHE["signature"] = None
        return jsonify({"id": arrival.id, "message": "创建成功"}), 201

    @app.put("/api/db-arrival/<int:record_id>")
    def db_arrival_update(record_id):
        arrival = db.session.get(ArrivalOrder, record_id)
        if not arrival:
            return jsonify({"message": "记录不存在"}), 404
        data = request.get_json() or {}
        if "sourceFile" in data: arrival.source_file = clean_text(data["sourceFile"])
        if "arrivalDate" in data: arrival.arrival_date = clean_text(data["arrivalDate"])
        if "contractNo" in data: arrival.contract_no = clean_text(data["contractNo"])
        if "customerCode" in data: arrival.customer_code = clean_text(data["customerCode"])
        if "sequence" in data: arrival.sequence = clean_text(data["sequence"])
        if "partNo" in data: arrival.part_no = clean_text(data["partNo"])
        if "interchangePartNo" in data: arrival.interchange_part_no = clean_text(data["interchangePartNo"])
        if "partName" in data: arrival.part_name = clean_text(data["partName"])
        if "quantity" in data: arrival.quantity = clean_number(data["quantity"])
        if "unitPrice" in data: arrival.unit_price = clean_number(data["unitPrice"])
        if "totalPrice" in data: arrival.total_price = clean_number(data["totalPrice"])
        db.session.commit()
        ARRIVAL_MATCH_CACHE["signature"] = None
        return jsonify({"message": "更新成功"})

    @app.delete("/api/db-arrival/<int:record_id>")
    def db_arrival_delete(record_id):
        arrival = db.session.get(ArrivalOrder, record_id)
        if not arrival:
            return jsonify({"message": "记录不存在"}), 404
        db.session.delete(arrival)
        db.session.commit()
        ARRIVAL_MATCH_CACHE["signature"] = None
        return jsonify({"message": "删除成功"})

    @app.get("/api/orders")
    def get_orders():
        customer = request.args.get("customer", "").strip()
        contract = request.args.get("contract", "").strip()
        part = request.args.get("part", "").strip()
        page = max(int(request.args.get("page", 1)), 1)
        page_size = min(max(int(request.args.get("pageSize", 20)), 1), 100)

        query = base_order_query()
        if customer:
            customer_clause = customer_filter_clause(customer)
            if customer_clause is not None:
                query = query.filter(customer_clause)
        if contract:
            query = query.filter(Order.contract_no.ilike(f"%{contract}%"))
        if part:
            query = query.filter(
                or_(
                    Order.part_no.ilike(f"%{part}%"),
                    Order.interchange_part_no.ilike(f"%{part}%"),
                )
            )

        total = query.count()
        items = (
            query.order_by(Order.id.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )

        return jsonify(
            {
                "items": [serialize_order(order) for order in items],
                "total": total,
                "page": page,
                "pageSize": page_size,
            }
        )

    def _query_contract_summary(search, year):
        query = ContractSummary.query
        if search:
            query = query.filter(or_(
                ContractSummary.contract_no.ilike(f"%{search}%"),
                ContractSummary.customer_code.ilike(f"%{search}%"),
            ))
        if year:
            query = query.filter(ContractSummary.order_year == year)
        return query.order_by(ContractSummary.total_amount.desc())

    def _contract_summary_to_dict(row):
        try:
            arrival_history = json.loads(row.arrival_history_json or "[]")
        except Exception:
            arrival_history = []
        return {
            "contractNo": row.contract_no,
            "customerCode": row.customer_code,
            "orderYear": row.order_year,
            "orderDate": row.order_date,
            "partItemCount": row.part_item_count,
            "totalQuantity": row.total_quantity,
            "totalAmount": row.total_amount,
            "arrivalQuantity": row.arrival_quantity,
            "arrivalAmount": row.arrival_amount,
            "notArrivalAmount": row.not_arrival_amount,
            "arrivalRatio": row.arrival_ratio,
            "arrivalStatus": row.arrival_status,
            "arrivalHistory": arrival_history,
        }

    @app.get("/api/contracts")
    def get_contracts():
        search = request.args.get("search", "").strip()
        year = request.args.get("year", "").strip()
        page = max(int(request.args.get("page", 1)), 1)
        page_size = min(max(int(request.args.get("pageSize", 20)), 1), 100)

        query = _query_contract_summary(search, year)
        total = query.count()
        items = query.offset((page - 1) * page_size).limit(page_size).all()

        available_years = [
            r[0] for r in db.session.query(ContractSummary.order_year)
            .filter(ContractSummary.order_year.isnot(None))
            .distinct().order_by(ContractSummary.order_year.desc()).all()
            if r[0]
        ]

        return jsonify({
            "items": [_contract_summary_to_dict(r) for r in items],
            "total": total,
            "page": page,
            "pageSize": page_size,
            "availableYears": available_years,
        })

    @app.get("/api/contracts/export")
    def export_contracts():
        search = request.args.get("search", "").strip()
        year = request.args.get("year", "").strip()

        rows = _query_contract_summary(search, year).all()
        export_rows = [
            {
                "序号": i + 1,
                "合同号": r.contract_no or "",
                "下单年份": r.order_year or "",
                "下单日期": format_order_date_text(r.order_date),
                "客户": r.customer_code or "",
                "零件总个数": r.total_quantity or 0,
                "合同总额": r.total_amount or 0,
                "到货状态": r.arrival_status or "未到货",
                "到货比例": f"{r.arrival_ratio or 0}%",
                "到货总个数": r.arrival_quantity or 0,
                "到货总金额": r.arrival_amount or 0,
                "未到货总金额": r.not_arrival_amount or 0,
            }
            for i, r in enumerate(rows)
        ]

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            pd.DataFrame(export_rows).to_excel(writer, sheet_name="contracts", index=False)
        output.seek(0)

        search_label = re.sub(r"[^\w\-一-龥]+", "_", search).strip("_")
        filename = f"合同信息_{year or '全部年份'}"
        if search_label:
            filename += f"_{search_label}"
        filename += "_全部结果.xlsx"

        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @app.get("/api/contract-items")
    def get_contract_items():
        contract_no = request.args.get("contractNo", "").strip()
        customer_code = request.args.get("customerCode", "").strip()

        query = base_order_query()
        if contract_no:
            query = query.filter(Order.contract_no == contract_no)
        if customer_code:
            customer_clause = customer_filter_clause(customer_code)
            if customer_clause is not None:
                query = query.filter(customer_clause)

        items = query.order_by(func.cast(Order.sequence, db.Integer).asc(), Order.id.asc()).all()
        arrival_index = get_arrival_match_index()

        results = []
        for order in items:
            order_data = serialize_order(order)
            customer_norm = _to_upper_text(normalize_customer_code(order.customer_code))
            contract_norm = _to_upper_text(order.contract_no)
            candidates = arrival_index.get((customer_norm, contract_norm), []) if customer_norm and contract_norm else []
            arrival_quantity, arrival_dates = _match_arrivals_for_order(order, candidates)
            arrival_details = _get_arrival_details_for_order(order, candidates)
            order_data["arrivalQuantity"] = arrival_quantity
            order_data["arrivalDates"] = arrival_dates
            order_data["arrivalDateText"] = "、".join(arrival_dates) if arrival_dates else ""
            order_data["arrivalDetails"] = arrival_details
            results.append(order_data)

        return jsonify(
            {
                "items": results,
                "total": len(items),
                "contractNo": contract_no,
                "customerCode": customer_code,
            }
        )

    @app.get("/api/customers")
    def get_customers():
        search = request.args.get("search", "").strip()
        page = max(int(request.args.get("page", 1)), 1)
        page_size = min(max(int(request.args.get("pageSize", 20)), 1), 100)

        query = CustomerSummary.query
        if search:
            query = query.filter(CustomerSummary.customer_code.ilike(f"%{search}%"))
        query = query.order_by(CustomerSummary.total_amount.desc())

        total = query.count()
        items = query.offset((page - 1) * page_size).limit(page_size).all()

        return jsonify({
            "items": [
                {
                    "customerCode": r.customer_code,
                    "contractCount": r.contract_count,
                    "orderCount": r.order_count,
                    "totalAmount": r.total_amount,
                    "latestContractTime": r.latest_contract_time,
                    "latestContractNo": r.latest_contract_no,
                }
                for r in items
            ],
            "total": total,
            "page": page,
            "pageSize": page_size,
        })

    @app.get("/api/customer-analysis")
    def get_customer_analysis():
        customer_code = request.args.get("customerCode", "").strip()
        if not customer_code:
            return jsonify({"summary": None, "topParts": [], "halfyearStats": {}, "quarterStats": {}, "yearlyStats": []})

        normalized_customer = normalize_customer_code(customer_code)
        rows = Order.query.filter(
            Order.contract_no.isnot(None),
            Order.customer_code == normalized_customer,
        ).all()
        if not rows:
            return jsonify({"summary": None, "topParts": [], "halfyearStats": {}, "quarterStats": {}, "yearlyStats": []})

        period_buckets = create_period_buckets()
        yearly_buckets = {}
        part_map = {}
        contract_set = set()
        total_amount = 0.0

        for row in rows:
            amount = float(row.total_price or 0)
            quantity = float(row.quantity or 0)
            total_amount += amount
            contract_set.add(row.contract_no)

            contract_token = row.contract_no or f"row-{row.id}"
            month = get_contract_month(row.contract_no)
            year = get_contract_year(row.contract_no)
            update_period_buckets(period_buckets, contract_token, month, amount, quantity)

            part_key = row.part_no or row.interchange_part_no or f"UNKNOWN-{row.id}"
            year_item = get_year_bucket(yearly_buckets, year)
            year_item["contracts"].add(row.contract_no)
            year_item["entryCount"] += 1
            year_item["totalAmount"] += amount
            year_item["totalQuantity"] += quantity
            if part_key and not str(part_key).startswith("UNKNOWN-"):
                year_item["partSet"].add(part_key)
            update_period_buckets(year_item["periods"], contract_token, month, amount, quantity)
            update_price_range(year_item, row.unit_price)

            year_part_item = year_item["partMap"].setdefault(
                part_key,
                {
                    "partNo": row.part_no,
                    "interchangePartNo": row.interchange_part_no,
                    "partName": row.part_name,
                    "entryCount": 0,
                    "contractSet": set(),
                    "totalQuantity": 0.0,
                    "totalAmount": 0.0,
                    "minUnitPrice": None,
                    "maxUnitPrice": None,
                },
            )
            year_part_item["entryCount"] += 1
            year_part_item["contractSet"].add(row.contract_no)
            year_part_item["totalQuantity"] += quantity
            year_part_item["totalAmount"] += amount
            if row.unit_price is not None:
                unit_price = float(row.unit_price)
                year_part_item["minUnitPrice"] = unit_price if year_part_item["minUnitPrice"] is None else min(year_part_item["minUnitPrice"], unit_price)
                year_part_item["maxUnitPrice"] = unit_price if year_part_item["maxUnitPrice"] is None else max(year_part_item["maxUnitPrice"], unit_price)

            part_item = part_map.setdefault(
                part_key,
                {
                    "partNo": row.part_no,
                    "interchangePartNo": row.interchange_part_no,
                    "partName": row.part_name,
                    "entryCount": 0,
                    "contractSet": set(),
                    "totalQuantity": 0.0,
                    "totalAmount": 0.0,
                    "minUnitPrice": None,
                    "maxUnitPrice": None,
                },
            )
            part_item["entryCount"] += 1
            part_item["contractSet"].add(row.contract_no)
            part_item["totalQuantity"] += quantity
            part_item["totalAmount"] += amount

            if row.unit_price is not None:
                unit_price = float(row.unit_price)
                part_item["minUnitPrice"] = unit_price if part_item["minUnitPrice"] is None else min(part_item["minUnitPrice"], unit_price)
                part_item["maxUnitPrice"] = unit_price if part_item["maxUnitPrice"] is None else max(part_item["maxUnitPrice"], unit_price)

        top_parts = []
        for item in part_map.values():
            top_parts.append(
                {
                    "partNo": normalize_part_code(item["partNo"]),
                    "interchangePartNo": normalize_part_code(item["interchangePartNo"]),
                    "partName": item["partName"],
                    "entryCount": item["entryCount"],
                    "contractCount": len({x for x in item["contractSet"] if x}),
                    "totalQuantity": round(item["totalQuantity"], 2),
                    "totalAmount": round(item["totalAmount"], 2),
                    "minUnitPrice": None if item["minUnitPrice"] is None else round(item["minUnitPrice"], 2),
                    "maxUnitPrice": None if item["maxUnitPrice"] is None else round(item["maxUnitPrice"], 2),
                }
            )

        top_parts.sort(key=lambda x: x["totalAmount"], reverse=True)
        finalized_periods = finalize_period_buckets(period_buckets)
        yearly_stats = finalize_yearly_buckets(yearly_buckets)

        return jsonify(
            {
                "summary": {
                    "customerCode": normalized_customer,
                    "contractCount": len({x for x in contract_set if x}),
                    "partTypeCount": len(part_map),
                    "totalAmount": round(total_amount, 2),
                    "partList": sorted([x for x in part_map.keys() if x and not str(x).startswith("UNKNOWN-")]),
                },
                "halfyearStats": {
                    "H1": finalized_periods["H1"],
                    "H2": finalized_periods["H2"],
                },
                "quarterStats": {
                    "Q1": finalized_periods["Q1"],
                    "Q2": finalized_periods["Q2"],
                    "Q3": finalized_periods["Q3"],
                    "Q4": finalized_periods["Q4"],
                },
                "yearlyStats": yearly_stats,
                "topParts": top_parts,
            }
        )

    @app.get("/api/parts")
    def get_parts():
        search = request.args.get("search", "").strip()
        page = max(int(request.args.get("page", 1)), 1)
        page_size = min(max(int(request.args.get("pageSize", 20)), 1), 100)

        query = PartSummary.query
        if search:
            query = query.filter(or_(
                PartSummary.part_no.ilike(f"%{search}%"),
                PartSummary.interchange_part_no.ilike(f"%{search}%"),
                PartSummary.part_name.ilike(f"%{search}%"),
            ))
        query = query.order_by(PartSummary.total_amount.desc())

        total = query.count()
        items = query.offset((page - 1) * page_size).limit(page_size).all()

        return jsonify({
            "items": [
                {
                    "partNo": r.part_no,
                    "interchangePartNo": r.interchange_part_no,
                    "partName": r.part_name,
                    "contractCount": r.contract_count,
                    "orderCount": r.order_count,
                    "totalQuantity": r.total_quantity,
                    "totalAmount": r.total_amount,
                    "minUnitPrice": r.min_unit_price,
                    "maxUnitPrice": r.max_unit_price,
                }
                for r in items
            ],
            "total": total,
            "page": page,
            "pageSize": page_size,
        }
        )

    @app.get("/api/part-analysis")
    def get_part_analysis():
        part_no = request.args.get("partNo", "").strip()
        interchange_part_no = request.args.get("interchangePartNo", "").strip()

        query = base_order_query()
        if part_no or interchange_part_no:
            keys = expand_part_lookup_keys([part_no, interchange_part_no])
            query = query.filter(
                or_(
                    Order.part_no.in_(keys),
                    Order.interchange_part_no.in_(keys),
                )
            )

        rows = query.all()
        if not rows:
            return jsonify({"summary": None, "halfyearStats": {}, "quarterStats": {}, "yearlyStats": [], "customerBreakdown": []})

        period_buckets = create_period_buckets()
        yearly_buckets = {}
        customer_map = {}
        total_amount = 0.0
        total_quantity = 0.0
        min_unit_price = None
        max_unit_price = None
        customer_set = set()

        main_part_no = next((row.part_no for row in rows if row.part_no), part_no or interchange_part_no)
        main_interchange = next((row.interchange_part_no for row in rows if row.interchange_part_no), interchange_part_no or None)
        main_part_name = next((row.part_name for row in rows if row.part_name), None)

        for row in rows:
            amount = float(row.total_price or 0)
            quantity = float(row.quantity or 0)
            total_amount += amount
            total_quantity += quantity

            contract_token = row.contract_no or f"row-{row.id}"
            month = get_contract_month(row.contract_no)
            year = get_contract_year(row.contract_no)
            update_period_buckets(period_buckets, contract_token, month, amount, quantity)

            year_item = get_year_bucket(yearly_buckets, year)
            year_item["contracts"].add(row.contract_no)
            year_item["entryCount"] += 1
            year_item["totalAmount"] += amount
            year_item["totalQuantity"] += quantity
            update_price_range(year_item, row.unit_price)

            if row.unit_price is not None:
                unit_price = float(row.unit_price)
                min_unit_price = unit_price if min_unit_price is None else min(min_unit_price, unit_price)
                max_unit_price = unit_price if max_unit_price is None else max(max_unit_price, unit_price)

            customer_key = normalize_customer_code(row.customer_code) or "未知"
            customer_set.add(customer_key)
            year_item["customerSet"].add(customer_key)
            update_period_buckets(year_item["periods"], contract_token, month, amount, quantity)

            year_customer_item = year_item["customerMap"].setdefault(
                customer_key,
                {
                    "customerCode": customer_key,
                    "contractSet": set(),
                    "entryCount": 0,
                    "totalQuantity": 0.0,
                    "totalAmount": 0.0,
                    "minUnitPrice": None,
                    "maxUnitPrice": None,
                },
            )
            year_customer_item["contractSet"].add(row.contract_no)
            year_customer_item["entryCount"] += 1
            year_customer_item["totalQuantity"] += quantity
            year_customer_item["totalAmount"] += amount
            if row.unit_price is not None:
                unit_price = float(row.unit_price)
                year_customer_item["minUnitPrice"] = unit_price if year_customer_item["minUnitPrice"] is None else min(year_customer_item["minUnitPrice"], unit_price)
                year_customer_item["maxUnitPrice"] = unit_price if year_customer_item["maxUnitPrice"] is None else max(year_customer_item["maxUnitPrice"], unit_price)

            customer_item = customer_map.setdefault(
                customer_key,
                {
                    "customerCode": customer_key,
                    "contractSet": set(),
                    "entryCount": 0,
                    "totalQuantity": 0.0,
                    "totalAmount": 0.0,
                    "minUnitPrice": None,
                    "maxUnitPrice": None,
                },
            )
            customer_item["contractSet"].add(row.contract_no)
            customer_item["entryCount"] += 1
            customer_item["totalQuantity"] += quantity
            customer_item["totalAmount"] += amount

            if row.unit_price is not None:
                unit_price = float(row.unit_price)
                customer_item["minUnitPrice"] = unit_price if customer_item["minUnitPrice"] is None else min(customer_item["minUnitPrice"], unit_price)
                customer_item["maxUnitPrice"] = unit_price if customer_item["maxUnitPrice"] is None else max(customer_item["maxUnitPrice"], unit_price)

        customer_breakdown = [
            {
                "customerCode": item["customerCode"],
                "contractCount": len({x for x in item["contractSet"] if x}),
                "entryCount": item["entryCount"],
                "totalQuantity": round(item["totalQuantity"], 2),
                "totalAmount": round(item["totalAmount"], 2),
                "minUnitPrice": None if item["minUnitPrice"] is None else round(item["minUnitPrice"], 2),
                "maxUnitPrice": None if item["maxUnitPrice"] is None else round(item["maxUnitPrice"], 2),
            }
            for item in customer_map.values()
        ]
        customer_breakdown.sort(key=lambda x: x["totalAmount"], reverse=True)

        finalized_periods = finalize_period_buckets(period_buckets)
        yearly_stats = finalize_yearly_buckets(yearly_buckets)

        return jsonify(
            {
                "summary": {
                    "partNo": normalize_part_code(main_part_no),
                    "interchangePartNo": normalize_part_code(main_interchange),
                    "partName": main_part_name,
                    "customerCount": len(customer_set),
                    "customerList": sorted(customer_set),
                    "totalQuantity": round(total_quantity, 2),
                    "totalAmount": round(total_amount, 2),
                    "minUnitPrice": None if min_unit_price is None else round(min_unit_price, 2),
                    "maxUnitPrice": None if max_unit_price is None else round(max_unit_price, 2),
                },
                "halfyearStats": {
                    "H1": finalized_periods["H1"],
                    "H2": finalized_periods["H2"],
                },
                "quarterStats": {
                    "Q1": finalized_periods["Q1"],
                    "Q2": finalized_periods["Q2"],
                    "Q3": finalized_periods["Q3"],
                    "Q4": finalized_periods["Q4"],
                },
                "yearlyStats": yearly_stats,
                "customerBreakdown": customer_breakdown,
            }
        )

    @app.post("/api/upload-orders")
    def upload_orders():
        if "file" not in request.files:
            return jsonify({"message": "未选择文件"}), 400
        file = request.files["file"]
        if not file.filename or not file.filename.lower().endswith(".xlsx"):
            return jsonify({"message": "请上传 .xlsx 文件"}), 400

        temp_path = BASE_DIR / "_upload_orders_tmp.xlsx"
        file.save(str(temp_path))

        result = sync_orders_from_excel(temp_path)
        if result.get("error"):
            temp_path.unlink(missing_ok=True)
            return jsonify({"message": result.get("message")}), 400

        temp_path.replace(SOURCE_FILE)
        ARRIVAL_MATCH_CACHE["signature"] = None
        ARRIVAL_MATCH_CACHE["index"] = defaultdict(list)
        rebuild_all_summaries()
        return jsonify(result)

    @app.post("/api/upload-arrivals")
    def upload_arrivals():
        files = request.files.getlist("files")
        if not files or all(not f.filename for f in files):
            return jsonify({"message": "未选择文件"}), 400

        existing_sources = {
            value
            for (value,) in db.session.query(ArrivalOrder.source_file)
            .filter(ArrivalOrder.source_file.isnot(None))
            .distinct()
            .all()
            if value
        }

        saved, skipped, errors = [], [], []
        if not ARRIVAL_DIR.exists():
            ARRIVAL_DIR.mkdir(parents=True)

        for file in files:
            name = file.filename or ""
            if not name.lower().endswith(".xlsx"):
                errors.append({"filename": name, "reason": "非 xlsx 文件"})
                continue
            if name in existing_sources:
                skipped.append(name)
                continue
            file.save(str(ARRIVAL_DIR / name))
            saved.append(name)

        if not saved:
            return jsonify({
                "message": "没有新文件需要导入（已跳过重复或无效文件）",
                "imported": 0,
                "savedFiles": [],
                "skippedFiles": skipped,
                "errorFiles": errors,
            })

        result = sync_arrivals_from_files()
        rebuild_all_summaries()
        result["savedFiles"] = saved
        result["skippedFiles"] = skipped
        result["errorFiles"] = errors
        return jsonify(result)

    @app.post("/api/sync")
    def sync_data():
        order_result = sync_orders_from_excel()
        arrival_result = sync_arrivals_from_files()
        rebuild_all_summaries()
        return jsonify({
            "orders": order_result,
            "arrivals": arrival_result,
            "message": "订单与到货数据已同步",
        })

    return app


app = create_app()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5001, debug=False)
